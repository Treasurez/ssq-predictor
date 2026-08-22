#!/usr/bin/env python3
"""
双色球彩票图片OCR识别 - RapidOCR版
=====================================
核心思路：RapidOCR全图识别 → 按"红球"/"蓝球"关键字位置分组 → 正则提取号码
RapidOCR是PaddleOCR的ONNX推理版，识别率相同，但无需安装paddlepaddle，无版本兼容问题。

安装依赖：
    pip3 install rapidocr-onnxruntime opencv-python numpy pandas openpyxl

运行：
    python3 lottery_paddle.py --image ./test.jpg --debug
    python3 lottery_paddle.py --folder ./lottery_img
"""
import re
import os
import sys
import argparse
from collections import defaultdict, Counter

import cv2
import numpy as np
import pandas as pd

# ===================== OCR初始化 =====================
_ocr_instance = None

def get_ocr():
    global _ocr_instance
    if _ocr_instance is None:
        from rapidocr_onnxruntime import RapidOCR
        _ocr_instance = RapidOCR()
    return _ocr_instance


def ocr_image(img):
    """对图片执行OCR，返回 [(text, cx, cy, w, h, prob), ...]
    RapidOCR输出: result = [[bbox, text, confidence], ...]
    """
    ocr = get_ocr()
    result, _ = ocr(img)
    items = []
    if not result:
        return items
    for line in result:
        try:
            bbox, text, prob = line
        except Exception:
            continue
        if not text or not text.strip():
            continue
        xs = [p[0] for p in bbox]
        ys = [p[1] for p in bbox]
        cx = sum(xs) / len(xs)
        cy = sum(ys) / len(ys)
        w = max(xs) - min(xs)
        h = max(ys) - min(ys)
        items.append((text.strip(), cx, cy, w, h, prob))
    return items


# ===================== 号码提取 =====================
def extract_nums(text, max_range=33):
    """从文本中提取1-max_range范围的数字，含OCR颠倒修正和连写拆分"""
    nums = []
    # 先移除倍数标记 [1倍] 等
    clean = re.sub(r'\[?\d+\s*[倍份服股倚伯侪倌倩宿佾伤府佰]\]?', ' ', text)
    for m in re.finditer(r'\d+', clean):
        raw = m.group()
        if len(raw) <= 2:
            n = int(raw)
            if 10 <= n <= 99:
                rev = int(str(n)[::-1])
                if (1 <= rev <= max_range) and not (1 <= n <= max_range):
                    n = rev
            if 1 <= n <= max_range:
                nums.append(n)
        else:
            # 连写拆分：每2位尝试
            i = 0
            while i < len(raw):
                seg = raw[i:i+2]
                if len(seg) == 2:
                    n = int(seg)
                    if 10 <= n <= 99:
                        rev = int(str(n)[::-1])
                        if (1 <= rev <= max_range) and not (1 <= n <= max_range):
                            n = rev
                    if 1 <= n <= max_range:
                        nums.append(n)
                        i += 2
                        continue
                # 1位
                n = int(raw[i])
                if 1 <= n <= max_range:
                    nums.append(n)
                i += 1
    return nums


def is_red_keyword(text):
    """检测是否为红球关键字行"""
    return bool(re.search(r'红[球胆拖爬皿昭粑色醯:：;；]', text))


def is_blue_keyword(text):
    """检测是否为蓝球关键字行"""
    return bool(re.search(r'蓝[球胆琛求撇搔监益盔盛盟:：;；]', text))


def is_noise(text):
    """检测是否为噪音文本（时间戳、电话、价格、装饰文字等）"""
    if re.search(r'\d+[:;]\d+[:;]\d+', text):  # 时间戳
        return True
    if re.search(r'\d+-\d+-\d+', text):  # 电话/序列号
        return True
    if re.search(r'20\d{2}[-/年]\d{1,2}', text):  # 日期
        return True
    if re.search(r'\d+\.\d+', text):  # 价格
        return True
    if '元' in text or '兀' in text:
        return True
    if '期' in text and not re.search(r'红|蓝', text):
        return True
    noise_words = ['财源', '滚滚', '公益', '感谢', '福利', '彩票', '福地', '暴富',
                   '必中', '团队', '祝福', '上岸', '打卡', '生肖', '双色球',
                   'CHAN', 'WELFARE', 'LOTTERY']
    for nw in noise_words:
        if nw in text:
            return True
    return False


# ===================== 核心解析逻辑 =====================
def parse_lottery_image(img_path, debug=False):
    """解析单张彩票图片，返回 [{red, blue}, ...]"""
    img = cv2.imread(img_path)
    if img is None:
        print(f"  无法读取图片: {img_path}")
        return []
    h, w = img.shape[:2]
    if debug:
        print(f"  图片尺寸: {w}x{h}")

    # ===== 主方案：关键字分组（红球识别率高，配对逻辑已重写）=====
    groups = parse_by_keyword_position(img, debug)

    if groups and len(groups) >= 8:
        if debug:
            print(f"  [关键字分组] 成功识别 {len(groups)} 组")
        return groups[:12]

    # ===== 备选：网格区域过滤 =====
    if debug:
        print(f"  [关键字分组] 仅 {len(groups) if groups else 0} 组，尝试网格过滤")
    groups2 = parse_by_grid_filter(img, debug)
    if groups2 and len(groups2) > len(groups or []):
        return groups2[:12]

    return (groups or [])[:12]


def detect_lottery_area(img):
    """红色像素检测彩票区域，返回 (x1, y1, x2, y2)"""
    h, w = img.shape[:2]
    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    lower1 = np.array([0, 100, 100]); upper1 = np.array([10, 255, 255])
    lower2 = np.array([170, 100, 100]); upper2 = np.array([179, 255, 255])
    red_mask = cv2.bitwise_or(cv2.inRange(hsv, lower1, upper1), cv2.inRange(hsv, lower2, upper2))
    row_ratio = np.sum(red_mask > 0, axis=1) / w
    col_ratio = np.sum(red_mask > 0, axis=0) / h

    def find_longest(ratio, threshold, merge_gap):
        is_low = ratio < threshold
        segs = []
        cur = None
        for i in range(len(ratio)):
            if is_low[i] and cur is None:
                cur = i
            elif not is_low[i] and cur is not None:
                segs.append((cur, i)); cur = None
        if cur is not None:
            segs.append((cur, len(ratio)))
        merged = []
        for s in segs:
            if merged and s[0] - merged[-1][1] < merge_gap:
                merged[-1] = (merged[-1][0], s[1])
            else:
                merged.append(s)
        return max(merged, key=lambda s: s[1]-s[0]) if merged else (0, len(ratio))

    y1, y2 = find_longest(row_ratio, 0.55, 60)
    x1, x2 = find_longest(col_ratio, 0.55, 40)
    return max(0, x1-10), max(0, y1-10), min(w, x2+10), min(h, y2+10)


def parse_by_grid_filter(img, debug=False):
    """主方案：全图OCR一次 → 红色像素检测区域 → 4×3网格过滤 → 每cell单独提取"""
    h, w = img.shape[:2]

    # 全图OCR（只做一次）
    items = ocr_image(img)
    if debug:
        print(f"  全图OCR识别到 {len(items)} 个文字块")

    # 过滤噪音
    items = [it for it in items if not is_noise(it[0])]
    if debug:
        print(f"  过滤噪音后 {len(items)} 个文字块")

    # 检测彩票区域
    x1, y1, x2, y2 = detect_lottery_area(img)
    if debug:
        print(f"  [网格] 彩票区域 X=[{x1},{x2}] Y=[{y1},{y2}]")

    rows, cols = 4, 3
    cell_w = (x2 - x1) / cols
    cell_h = (y2 - y1) / rows
    margin = 5  # 小margin，避免切掉边界文字

    groups = []
    for r in range(rows):
        for c in range(cols):
            cx1 = x1 + c * cell_w - margin
            cy1 = y1 + r * cell_h - margin
            cx2 = x1 + (c+1) * cell_w + margin
            cy2 = y1 + (r+1) * cell_h + margin

            # 过滤落在当前cell内的文字块（中心点在cell内）
            cell_items = []
            for text, tx, ty, tw, th, prob in items:
                if cx1 <= tx <= cx2 and cy1 <= ty <= cy2:
                    cell_items.append((text, tx, ty, tw, th, prob))

            if debug:
                print(f"    格[{r},{c}] X=[{cx1:.0f},{cx2:.0f}] Y=[{cy1:.0f},{cy2:.0f}]: {len(cell_items)}个文字块")

            # 在cell内提取红球/蓝球
            # 策略：收集所有数字→按Y排序→找最大间距分割红蓝→上红下蓝
            all_nums = []  # [(num, cy), ...]
            for text, tx, ty, tw, th, prob in cell_items:
                ns = extract_nums(text, 33)
                for n in ns:
                    all_nums.append((n, ty))

            red_nums = []
            blue_nums = []

            if all_nums:
                all_nums.sort(key=lambda x: x[1])  # 按Y从小到大排序
                ys = [y for _, y in all_nums]

                # 找最大Y间距作为红蓝分割点
                max_gap = 0
                split_idx = len(all_nums) // 2
                for i in range(1, len(all_nums)):
                    gap = ys[i] - ys[i-1]
                    if gap > max_gap:
                        max_gap = gap
                        split_idx = i

                # 上半部分=红球，下半部分=蓝球（蓝球只取1-16）
                for n, _ in all_nums[:split_idx]:
                    if n not in red_nums:
                        red_nums.append(n)
                for n, _ in all_nums[split_idx:]:
                    if 1 <= n <= 16 and n not in blue_nums:
                        blue_nums.append(n)
                    elif n > 16 and n not in red_nums:
                        # 下半部分出现>16的数字，说明分割点偏上，补到红球
                        red_nums.append(n)

                # 如果蓝球为空，从红球中取1-16范围、Y坐标较大的作为蓝球
                if not blue_nums:
                    candidates = [(n, y) for n, y in all_nums if 1 <= n <= 16]
                    if candidates:
                        candidates.sort(key=lambda x: -x[1])  # Y大的优先（更靠下）
                        for n, _ in candidates[:3]:
                            if n not in blue_nums:
                                blue_nums.append(n)
                            if n in red_nums:
                                red_nums.remove(n)

                # 如果红球不足3个，把蓝球中>16的移回来（不应该发生）
                if len(red_nums) < 3:
                    for n in list(blue_nums):
                        if n > 16:
                            blue_nums.remove(n)
                            red_nums.append(n)

            # 降低门槛：红球≥3就输出，不跳过
            if len(red_nums) >= 3:
                red_final = sorted(set(red_nums))
                blue_final = sorted(set(blue_nums)) if blue_nums else []
                # 限制数量：红球最多8，蓝球最多4
                if len(red_final) > 8:
                    red_final = red_final[:8]
                if len(blue_final) > 4:
                    blue_final = blue_final[:4]
                # 蓝球为空时给个默认值，避免整组被过滤
                if not blue_final:
                    blue_final = [red_final[0]] if red_final and red_final[0] <= 16 else [1]
                groups.append({"red": red_final, "blue": blue_final})
                if debug:
                    print(f"      → 红{red_final} 蓝{blue_final}")
            else:
                if debug:
                    print(f"      → 跳过(红仅{len(red_nums)}个: {red_nums})")

    return groups


def parse_by_keyword_position(img, debug=False):
    """全图OCR → 红球行按Y聚类4行、每行按X排序3列 → 蓝球按X+Y就近配对"""
    h, w = img.shape[:2]
    items = ocr_image(img)
    if debug:
        print(f"  OCR识别到 {len(items)} 个文字块")

    # 过滤噪音
    items = [it for it in items if not is_noise(it[0])]
    if debug:
        print(f"  过滤噪音后 {len(items)} 个文字块")

    # 收集红球行、蓝球行、其他数字行（含cx, cy）
    red_lines = []    # [(cx, cy, [nums]), ...]
    blue_lines = []   # [(cx, cy, [nums]), ...]
    other_num_lines = []  # [(cx, cy, [nums]), ...]

    for text, cx, cy, bw, bh, prob in items:
        if is_red_keyword(text):
            nums = extract_nums(text, 33)
            if nums:
                red_lines.append((cx, cy, nums))
                if debug:
                    print(f"    红球行 cx={cx:.0f} cy={cy:.0f}: {nums} (原文: {text})")
        elif is_blue_keyword(text):
            nums = extract_nums(text, 16)
            if nums:
                blue_lines.append((cx, cy, nums))
                if debug:
                    print(f"    蓝球行 cx={cx:.0f} cy={cy:.0f}: {nums} (原文: {text})")
        else:
            nums = extract_nums(text, 33)
            if nums:
                other_num_lines.append((cx, cy, nums))

    if not red_lines:
        if debug:
            print("  未识别到红球关键字")
        return []

    # ===== 红球行按Y聚类成4行，每行内按X排序 =====
    red_lines.sort(key=lambda x: x[1])  # 先按Y排序
    rows = []
    current_row = [red_lines[0]]
    for i in range(1, len(red_lines)):
        if red_lines[i][1] - red_lines[i-1][1] < h * 0.08:  # Y差距小于图片高度8%算同一行
            current_row.append(red_lines[i])
        else:
            rows.append(current_row)
            current_row = [red_lines[i]]
    rows.append(current_row)

    if debug:
        print(f"  红球行聚类成 {len(rows)} 行: {[len(r) for r in rows]}")

    # 每行内按X排序，然后按行优先展开
    ordered_red = []
    for row in rows:
        row.sort(key=lambda x: x[0])  # 按X排序
        ordered_red.extend(row)

    if debug:
        print(f"  排序后红球行顺序: {[(f'cx={cx:.0f}', f'cy={cy:.0f}') for cx, cy, _ in ordered_red]}")

    # ===== 每个红球行配对蓝球行（X接近 + Y在下方）=====
    groups = []
    used_blue = set()

    for idx, (r_cx, r_cy, red_nums) in enumerate(ordered_red):
        # 找X接近、Y在红球行下方的蓝球关键字行
        best_blue = None
        best_score = float('inf')
        for j, (b_cx, b_cy, b_nums) in enumerate(blue_lines):
            if j in used_blue:
                continue
            if b_cy >= r_cy - 20:
                dx = abs(b_cx - r_cx)
                dy = b_cy - r_cy
                score = dx * 2 + dy
                if score < best_score and dx < w * 0.35:
                    best_score = score
                    best_blue = j

        blue_nums = []
        if best_blue is not None:
            used_blue.add(best_blue)
            blue_nums = list(blue_lines[best_blue][2])
            if debug:
                print(f"    组{idx+1} 红球cx={r_cx:.0f} → 配对蓝球cx={blue_lines[best_blue][0]:.0f}: {blue_nums}")
        else:
            if debug:
                print(f"    组{idx+1} 红球cx={r_cx:.0f} → 未配对蓝球关键字行")

        # 【改进】从红球下方、X接近的数字行收集1-16范围的数字作为蓝球候选
        # Y范围缩小到70px，避免收集到相邻彩票的蓝球
        blue_candidates = Counter()
        blue_keyword_nums = set()  # 蓝球关键字行中的数字（优先级最高）
        if best_blue is not None:
            for n in blue_lines[best_blue][2]:
                blue_keyword_nums.add(n)
        for cx, cy, nums in other_num_lines:
            if r_cy - 5 < cy < r_cy + 70 and abs(cx - r_cx) < w * 0.30:
                for n in nums:
                    if 1 <= n <= 16:
                        blue_candidates[n] += 1

        # 蓝球 = 关键字行数字 + 候选数字（去重）
        for n in blue_keyword_nums:
            if n not in blue_nums:
                blue_nums.append(n)
        for n in blue_candidates:
            if n not in blue_nums:
                blue_nums.append(n)

        # 【二次识别】红球关键字行不足5个时，对该区域多种预处理融合OCR
        if len(red_nums) < 5:
            h_img, w_img = img.shape[:2]
            cx1 = max(0, int(r_cx - 130))
            cy1 = max(0, int(r_cy - 30))
            cx2 = min(w_img, int(r_cx + 130))
            cy2 = min(h_img, int(r_cy + 70))
            if cx2 - cx1 > 50 and cy2 - cy1 > 30:
                sub_img = img[cy1:cy2, cx1:cx2]
                # 3种预处理融合：原图、对比度增强、二值化
                variants = []
                # 1. 原图放大8倍
                variants.append(cv2.resize(sub_img, None, fx=8.0, fy=8.0, interpolation=cv2.INTER_CUBIC))
                # 2. 灰度+CLAHE对比度增强
                if len(sub_img.shape) == 3:
                    gray = cv2.cvtColor(sub_img, cv2.COLOR_BGR2GRAY)
                else:
                    gray = sub_img
                clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
                enhanced = cv2.cvtColor(clahe.apply(gray), cv2.COLOR_GRAY2BGR)
                variants.append(cv2.resize(enhanced, None, fx=8.0, fy=8.0, interpolation=cv2.INTER_CUBIC))
                # 3. 二值化
                _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                binary_bgr = cv2.cvtColor(binary, cv2.COLOR_GRAY2BGR)
                variants.append(cv2.resize(binary_bgr, None, fx=8.0, fy=8.0, interpolation=cv2.INTER_CUBIC))
                # 融合所有识别结果
                all_sub_items = []
                for v in variants:
                    all_sub_items.extend(ocr_image(v))
                # 从融合结果中提取：红球补充所有数字（包括1-16），蓝球补充在候选中的
                for text, tx, ty, tw, th, prob in all_sub_items:
                    if is_noise(text):
                        continue
                    if is_red_keyword(text):
                        for n in extract_nums(text, 33):
                            if n not in red_nums:
                                red_nums.append(n)
                    elif is_blue_keyword(text):
                        for n in extract_nums(text, 16):
                            if n not in blue_nums:
                                if n in blue_candidates or n in blue_keyword_nums:
                                    blue_nums.append(n)
                                    blue_keyword_nums.add(n)
                    else:
                        for n in extract_nums(text, 33):
                            if n > 16 and n not in red_nums:
                                red_nums.append(n)
                if debug:
                    print(f"    [二次识别] 区域[{cx1}:{cx2},{cy1}:{cy2}] 3种预处理融合 红球补充后: {sorted(red_nums)} 蓝球: {sorted(blue_nums)}")

        # 红球补充：从X接近、Y接近的其他数字行补充
        # 优先补充>16的数字（红球特征），范围严格避免混组
        if len(red_nums) < 7:
            for cx, cy, nums in other_num_lines:
                if abs(cy - r_cy) < 12 and abs(cx - r_cx) < w * 0.20:
                    for n in nums:
                        if n > 16 and n not in red_nums and 1 <= n <= 33:
                            red_nums.append(n)
        # >16的补充完还不够，再补充1-16的（最多补充3个）
        if len(red_nums) < 6:
            count = 0
            for cx, cy, nums in other_num_lines:
                if abs(cy - r_cy) < 12 and abs(cx - r_cx) < w * 0.20:
                    for n in nums:
                        if 1 <= n <= 16 and n not in red_nums and n not in blue_keyword_nums:
                            red_nums.append(n)
                            count += 1
                            if count >= 3:
                                break
                    if count >= 3:
                        break

        # 红球中1-16的数字如果漏识别了，从蓝球候选中"借回"
        # 限制：红球不足6个才借回，最多借回(6-当前数量)个
        # 优先借回频次=1的（频次>=2的更可能是蓝球），且不在蓝球关键字行中
        if len(red_nums) < 6:
            borrow_count = min(2, 6 - len(red_nums))
            borrow_candidates = [(n, f) for n, f in blue_candidates.items()
                                 if 1 <= n <= 16 and n not in red_nums
                                 and n not in blue_keyword_nums and f == 1]
            borrow_candidates.sort(key=lambda x: -x[1])
            for n, f in borrow_candidates[:borrow_count]:
                red_nums.append(n)

        # 后处理：红球超过7个时，移除1-16中蓝球候选频次最高的（更可能是蓝球）
        if len(red_nums) > 7:
            red_1_16 = [(n, blue_candidates.get(n, 0)) for n in red_nums if 1 <= n <= 16]
            if red_1_16:
                # 按频次降序、数字降序排序，移除第一个
                red_1_16.sort(key=lambda x: (-x[1], -x[0]))
                remove_n = red_1_16[0][0]
                red_nums.remove(remove_n)
                if debug:
                    print(f"    [后处理] 红球超过7个，移除{remove_n}（蓝球候选频次{blue_candidates.get(remove_n, 0)}）")

        # 蓝球不足3个时，从候选中补充（频次>=1都补，最多补到4个）
        if len(blue_nums) < 3:
            for n, f in sorted(blue_candidates.items(), key=lambda x: -x[1]):
                if 1 <= n <= 16 and n not in blue_nums:
                    blue_nums.append(n)
                    if len(blue_nums) >= 4:
                        break

        if len(red_nums) >= 3:
            red_final = sorted(set(red_nums))
            blue_final = sorted(set(blue_nums)) if blue_nums else []
            # 红球最多8个
            if len(red_final) > 8:
                red_final = red_final[:8]
            # 蓝球筛选：以关键字行数量为准，兼容1-4个蓝球的复式票
            keyword_in_final = [n for n in blue_final if n in blue_keyword_nums]
            other_in_final = [n for n in blue_final if n not in blue_keyword_nums]
            other_in_final.sort(key=lambda n: (-blue_candidates.get(n, 0), 1 if n in red_nums else 0))
            if len(keyword_in_final) >= 3:
                # 关键字行有3个以上，以关键字行数量为准（最多4个）
                target = min(len(keyword_in_final), 4)
                blue_final = (keyword_in_final + other_in_final)[:target]
            else:
                # 关键字行不足3个，默认补到3个（兼容大多数3蓝球复式票）
                blue_final = (keyword_in_final + other_in_final)[:3]
            blue_final = sorted(set(blue_final))
            if not blue_final:
                blue_final = [red_final[0]] if red_final and red_final[0] <= 16 else [1]
            groups.append({"red": red_final, "blue": blue_final})
            if debug:
                print(f"      → 红{red_final} 蓝{blue_final} (蓝球关键字: {sorted(blue_keyword_nums)}, 候选: {dict(blue_candidates)})")

    return groups


def parse_by_grid_crop(img, debug=False):
    """方案2：红色像素检测区域 + 4×3等分 + 每个子图单独OCR"""
    h, w = img.shape[:2]

    # 红色像素检测彩票区域
    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    lower1 = np.array([0, 100, 100]); upper1 = np.array([10, 255, 255])
    lower2 = np.array([170, 100, 100]); upper2 = np.array([179, 255, 255])
    red_mask = cv2.bitwise_or(cv2.inRange(hsv, lower1, upper1), cv2.inRange(hsv, lower2, upper2))
    row_ratio = np.sum(red_mask > 0, axis=1) / w
    col_ratio = np.sum(red_mask > 0, axis=0) / h

    def find_longest(ratio, threshold, merge_gap):
        is_low = ratio < threshold
        segs = []
        cur = None
        for i in range(len(ratio)):
            if is_low[i] and cur is None:
                cur = i
            elif not is_low[i] and cur is not None:
                segs.append((cur, i)); cur = None
        if cur is not None:
            segs.append((cur, len(ratio)))
        merged = []
        for s in segs:
            if merged and s[0] - merged[-1][1] < merge_gap:
                merged[-1] = (merged[-1][0], s[1])
            else:
                merged.append(s)
        return max(merged, key=lambda s: s[1]-s[0]) if merged else (0, len(ratio))

    y1, y2 = find_longest(row_ratio, 0.55, 60)
    x1, x2 = find_longest(col_ratio, 0.55, 40)
    y1, y2 = max(0, y1-10), min(h, y2+10)
    x1, x2 = max(0, x1-10), min(w, x2+10)

    if debug:
        print(f"  [网格] 彩票区域 X=[{x1},{x2}] Y=[{y1},{y2}]")

    rows, cols = 4, 3
    cell_w = (x2 - x1) / cols
    cell_h = (y2 - y1) / rows
    margin = 8

    groups = []
    for r in range(rows):
        for c in range(cols):
            cx1 = max(0, int(x1 + c * cell_w) + margin)
            cy1 = max(0, int(y1 + r * cell_h) + margin)
            cx2 = min(w, int(x1 + (c+1) * cell_w) - margin)
            cy2 = min(h, int(y1 + (r+1) * cell_h) - margin)
            if cx2 <= cx1 + 20 or cy2 <= cy1 + 20:
                continue

            sub = img[cy1:cy2, cx1:cx2]
            # 放大3倍
            sub_big = cv2.resize(sub, None, fx=3.0, fy=3.0, interpolation=cv2.INTER_CUBIC)
            items = ocr_image(sub_big)

            red_nums = []
            blue_nums = []
            for text, tx, ty, tw, th, prob in items:
                if is_noise(text):
                    continue
                if is_red_keyword(text):
                    red_nums.extend(extract_nums(text, 33))
                elif is_blue_keyword(text):
                    blue_nums.extend(extract_nums(text, 16))

            # 关键字没找到时，用Y位置分离（上半红，下半蓝）
            if not red_nums or not blue_nums:
                all_nums = []
                for text, tx, ty, tw, th, prob in items:
                    if is_noise(text):
                        continue
                    ns = extract_nums(text, 33)
                    for n in ns:
                        all_nums.append((n, ty))
                if all_nums:
                    all_nums.sort(key=lambda x: x[1])
                    ys = [y for _, y in all_nums]
                    # 找最大Y间距分割
                    max_gap = 0
                    split_idx = len(all_nums) // 2
                    for i in range(1, len(all_nums)):
                        gap = ys[i] - ys[i-1]
                        if gap > max_gap:
                            max_gap = gap
                            split_idx = i
                    if max_gap > 20:
                        for n, _ in all_nums[:split_idx]:
                            if n not in red_nums:
                                red_nums.append(n)
                        for n, _ in all_nums[split_idx:]:
                            if 1 <= n <= 16 and n not in blue_nums:
                                blue_nums.append(n)

            if len(red_nums) >= 4 and len(blue_nums) >= 1:
                red_final = sorted(set(red_nums))[:8]
                blue_final = sorted(set(blue_nums))[:4]
                groups.append({"red": red_final, "blue": blue_final})
                if debug:
                    print(f"    格[{r},{c}]: 红{red_final} 蓝{blue_final}")

    return groups


# ===================== 批量处理 & 导出 =====================
def batch_parse(folder_path, debug=False):
    total_groups = []
    for fname in sorted(os.listdir(folder_path)):
        if fname.lower().endswith(('.jpg', '.jpeg', '.png')):
            fpath = os.path.join(folder_path, fname)
            print(f"\n处理: {fname}")
            groups = parse_lottery_image(fpath, debug=debug)
            for g in groups:
                g["img_name"] = fname
            total_groups.extend(groups)
            print(f"  识别到 {len(groups)} 组")
    print(f"\n总计 {len(total_groups)} 组")
    return total_groups


def export_excel(groups, save_path="lottery_summary.xlsx"):
    data = []
    for i, g in enumerate(groups, 1):
        data.append({
            "序号": i,
            "红球": " ".join(f"{x:02d}" for x in g["red"]),
            "蓝球": " ".join(f"{x:02d}" for x in g["blue"]),
        })
    df = pd.DataFrame(data)
    df.to_excel(save_path, index=False, engine="openpyxl")

    # 样式
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        wb = load_workbook(save_path)
        ws = wb.active
        hf = Font(name="微软雅黑", size=12, bold=True)
        ha = Alignment(horizontal="center", vertical="center")
        df_font = Font(name="微软雅黑", size=11)
        da = Alignment(horizontal="center", vertical="center")
        bd = Border(*[Side(style="thin", color="000000")]*4)
        for cell in ws[1]:
            cell.font = hf; cell.alignment = ha; cell.border = bd
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.font = df_font; cell.alignment = da; cell.border = bd
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 32
        ws.column_dimensions["C"].width = 18
        wb.save(save_path)
    except Exception as e:
        print(f"  样式优化跳过: {e}")
    print(f"已导出: {save_path} ({len(data)}组)")


# ===================== 主入口 =====================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="双色球彩票OCR识别 (PaddleOCR版)")
    parser.add_argument("--folder", "-f", default="../lottery_img", help="图片文件夹")
    parser.add_argument("--image", "-i", help="单张图片路径")
    parser.add_argument("--output", "-o", default="lottery_summary.xlsx", help="输出Excel路径")
    parser.add_argument("--debug", "-d", action="store_true", help="调试模式")
    args = parser.parse_args()

    if args.image:
        groups = parse_lottery_image(args.image, debug=args.debug)
        for i, g in enumerate(groups, 1):
            print(f"  组{i}: 红{g['red']} 蓝{g['blue']}")
        if groups:
            export_excel(groups, args.output)
    else:
        groups = batch_parse(args.folder, debug=args.debug)
        if groups:
            export_excel(groups, args.output)
