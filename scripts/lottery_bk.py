import easyocr
import cv2
import re
import os
import sys
from collections import defaultdict
import numpy as np
import pandas as pd

# 导入全局配置（确保能找到同目录的 config 模块）
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import LOTTERY_SUMMARY_PATH

# ===================== 1. 模型检查与初始化 =====================
# 优先使用项目目录的模型，否则使用 EasyOCR 默认路径
PROJECT_MODEL_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'easyocr_models')
DEFAULT_MODEL_DIR = os.path.expanduser('~/.EasyOCR/model')

REQUIRED_MODELS = ['craft_mlt_25k.pth', 'zh_sim_g2.pth', 'english_g2.pth']

def check_models(model_dir):
    missing = []
    if not os.path.exists(model_dir):
        return REQUIRED_MODELS
    for model_name in REQUIRED_MODELS:
        model_path = os.path.join(model_dir, model_name)
        if not os.path.exists(model_path):
            missing.append(model_name)
        elif os.path.getsize(model_path) < 500000:
            missing.append(f"{model_name} (文件太小)")
    return missing

# 先检查项目目录
missing_in_project = check_models(PROJECT_MODEL_DIR)
if not missing_in_project:
    MODEL_DIR = PROJECT_MODEL_DIR
    print(f"使用项目目录的 EasyOCR 模型: {MODEL_DIR}")
else:
    # 检查默认目录
    missing_in_default = check_models(DEFAULT_MODEL_DIR)
    if not missing_in_default:
        MODEL_DIR = DEFAULT_MODEL_DIR
        print(f"使用 EasyOCR 默认目录的模型: {MODEL_DIR}")
    else:
        # 两个目录都不完整，尝试使用默认路径触发下载
        print("检测到缺少 EasyOCR 模型文件...")
        print(f"\n项目目录 {PROJECT_MODEL_DIR} 缺少: {missing_in_project}")
        print(f"默认目录 {DEFAULT_MODEL_DIR} 缺少: {missing_in_default}")
        print("\n正在尝试通过 EasyOCR 自动下载模型...")
        print("（首次下载可能需要几分钟，取决于网络速度）\n")
        
        try:
            reader = easyocr.Reader(
                ['ch_sim','en'],
                gpu=False,
                download_enabled=True,
                verbose=True
            )
            MODEL_DIR = DEFAULT_MODEL_DIR
            print("\n模型下载完成！")
        except Exception as e:
            print(f"\n自动下载失败：{e}")
            print("\n请手动下载模型文件：")
            print("  python3 scripts/download_easyocr_models.py")
            sys.exit(1)
        
        # 下载完成后重新检查
        missing_after = check_models(DEFAULT_MODEL_DIR)
        if missing_after:
            print(f"\n下载后仍缺少: {missing_after}")
            sys.exit(1)

# 创建 Reader
reader = easyocr.Reader(
    ['ch_sim','en'],
    gpu=False,
    model_storage_directory=MODEL_DIR,
    download_enabled=False,
    verbose=False
)
print("EasyOCR 模型加载成功！")


# ===================== 2. 图片预处理：多种增强策略 =====================
def load_image(img_path):
    """加载图片，直接返回原图（BGR格式）"""
    img = cv2.imread(img_path)
    if img is None:
        return None
    return img

def preprocess_images(img):
    """生成多种预处理版本的图片，用于多次尝试OCR
    
    返回: 列表 [(名称, 图片), ...]
    """
    variants = []
    
    # 1. 原图
    variants.append(("原图", img))
    
    # 2. 放大1.5倍 (小图片号码字体小，放大有助于OCR)
    h, w = img.shape[:2]
    if max(h, w) < 1500:
        scale = 1.5
        big = cv2.resize(img, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
        variants.append(("放大1.5x", big))
    
    # 3. 灰度图
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    variants.append(("灰度图", gray))
    
    # 4. 灰度+放大
    if max(h, w) < 1500:
        gray_big = cv2.resize(gray, None, fx=1.5, fy=1.5, interpolation=cv2.INTER_CUBIC)
        variants.append(("灰度+放大", gray_big))
    
    # 5. 对比度增强 (CLAHE)
    lab = cv2.cvtColor(img, cv2.COLOR_BGR2LAB)
    l, a, b = cv2.split(lab)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    l = clahe.apply(l)
    enhanced = cv2.merge([l, a, b])
    enhanced = cv2.cvtColor(enhanced, cv2.COLOR_LAB2BGR)
    variants.append(("对比度增强", enhanced))
    
    # 6. 二值化 (OTSU自动阈值)
    _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    variants.append(("二值化", binary))
    
    # 7. 反色二值化 (深色背景时)
    _, binary_inv = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    variants.append(("反色二值化", binary_inv))
    
    # 8. 自适应二值化
    adaptive = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
                                      cv2.THRESH_BINARY, 11, 2)
    variants.append(("自适应二值化", adaptive))
    
    return variants


def ocr_with_detail(img_variant):
    """对单张图片执行OCR，返回带detail=1的结果 (带位置和置信度)"""
    name, img = img_variant
    try:
        results = reader.readtext(img, detail=1)
        # 统一坐标缩放回原图尺寸（对于放大的图，坐标要除回去）
        scale = 1.0
        if "放大" in name:
            scale = 1.5
        
        normalized = []
        for bbox, text, prob in results:
            if scale != 1.0:
                scaled_bbox = [[p[0]/scale, p[1]/scale] for p in bbox]
            else:
                scaled_bbox = bbox
            normalized.append((scaled_bbox, text, prob))
        return name, normalized
    except Exception as e:
        print(f"    [{name}] OCR失败: {e}")
        return name, []


def _is_noise_ocr_text(text):
    """检测OCR文本是否为噪音（时间戳、电话号码、价格、装饰文字等非彩票号码）
    
    这类文本中的数字不应被提取为彩票号码：
    - 时间戳：09:39:55, 0:09;59, 09;39;42 等
    - 电话/序列号：2817-1204-5550, 431107487-001 等
    - 日期：2026-08-11, 20260921 等
    - 价格：15.12元, 42元, 42兀 等
    - 期号：048期, 期中5+2 等
    - 装饰文字：财源滚滚, 公益, 感谢 等
    - 福彩相关文字
    """
    if not text or not text.strip():
        return True

    text = text.strip()

    # 时间戳模式：数字:数字:数字 或 数字;数字;数字（如 09:39:55, 0:09;59）
    if re.search(r'\d+[:;]\d+[:;]\d+', text):
        return True

    # 电话/序列号模式：数字-数字-数字（如 2817-1204-5550, 431107487-001）
    if re.search(r'\d+-\d+-\d+', text):
        return True
    # 序列号：数字-数字（如 656-706, 80697-104）
    if re.search(r'\d{3,}-\d{2,}', text):
        return True

    # 日期模式：2026-08, 2026/08, 20260921（8位连续日期）
    if re.search(r'20\d{2}[-/年]\d{1,2}', text):
        return True
    if re.search(r'^20\d{6}$', text):
        return True

    # 价格模式：15.12元, 42元, 42兀, 5.12兀
    if re.search(r'\d+\.\d+', text):
        return True
    if '元' in text or '兀' in text:
        return True

    # 期号模式：048期, 期中, 期$
    if '期' in text:
        return True

    # 金额/注数相关
    if '注' in text or '股' in text:
        return True
    # 总注数/总股数元数据（如 "共6服 找6服", "共6股"）
    if '共' in text and ('服' in text or '份' in text or '股' in text):
        return True
    # 含"找"字的行（如"找6服"是找零信息）
    if '找' in text:
        return True

    # 装饰文字
    noise_words = ['财源', '滚滚', '公益', '感谢', '福利', '彩票', '醺', '骠',
                   '福地', '暴富', '必中', '团队', '祝福', '上岸', '打卡',
                   '中国', '中 国', '生肖', '双色球', '双色 球',
                   'CHAN', 'CHANES', 'GAON', 'GAOA', 'MLUA', 'MLUU',
                   'LTTIAY', 'LATTIAY', 'UILIAN', 'WULIAN', 'WUTAN',
                   'ITITAN', 'ITITANY', 'ULUIAN', 'WLIAA', 'MTITAN',
                   'T7-EBKL', 'CAg0', 'F7 EBAE', 'CASI', 'EBME-CASU']
    for nw in noise_words:
        if nw in text:
            return True

    # 纯英文字母串（如 FBOF, VA, CMINA 等OCR噪音）
    if re.search(r'^[A-Za-z\s\.\-\d]{0,5}[A-Za-z]{3,}[A-Za-z\s\.\-\d]*$', text) and not re.search(r'\d{2}', text):
        return True

    # 含有冒号但不是红球/蓝球关键字的（如 m;3:59, n潞什钢:3:]）
    if re.search(r'\d+[:;]\d+', text) and not any(kw in text for kw in ['红', '蓝', '球', '胆', '拖']):
        return True

    # "今晚" 等文字
    if '今晚' in text or '今' in text:
        return True

    return False


def merge_ocr_results(all_ocr_results):
    """合并多种预处理的OCR结果，去重并择优
    
    all_ocr_results: [(预处理名, [(bbox, text, prob), ...]), ...]
    返回:
        lines_text: 去重后的纯文本行列表（用于传统行解析）
        all_items: 带位置的数字/关键字列表 [{type, text, num, x, y, prob, source}, ...]
    """
    # 收集所有数字和关键字
    all_items_raw = []  # 先收集全部候选，再做位置去重（保留高置信度）
    
    for src_name, ocr_results in all_ocr_results:
        for bbox, text, prob in ocr_results:
            if not text or not text.strip():
                continue
            # 计算中心点
            xs = [p[0] for p in bbox]
            ys = [p[1] for p in bbox]
            cx, cy = sum(xs)/len(xs), sum(ys)/len(ys)
            w = max(xs) - min(xs)
            h = max(ys) - min(ys)
            all_items_raw.append({
                "text": text.strip(),
                "x": cx,
                "y": cy,
                "w": w,
                "h": h,
                "prob": prob,
                "source": src_name,
            })
    
    # ===== 改进去重算法：位置接近则只保留置信度最高的 =====
    # 对于叠印严重的图片（如1896d7b3），同一位置会出现3-5次重复识别，
    # 使用 cx/15, cy/15 粒度去重（比之前更激进），并保留置信度最高的
    bucket_dict = {}
    for it in all_items_raw:
        cx, cy = it["x"], it["y"]
        # 更粗的位置桶：20px * 20px 粒度合并（避免同一文字被多次识别）
        key = (round(cx/20), round(cy/20), it["text"])
        if key not in bucket_dict or it["prob"] > bucket_dict[key]["prob"]:
            bucket_dict[key] = it
    # 如果上面去重后还是太多（>400），进一步用更粗的粒度
    deduplicated = list(bucket_dict.values())
    if len(deduplicated) > 400:
        # 极端叠印：再用 30*30 粒度，相同文本不同位置也合并（只保留前3次）
        bucket2 = {}
        text_count = {}
        for it in deduplicated:
            k2 = (round(it["x"]/30), round(it["y"]/30), it["text"])
            if k2 not in bucket2:
                bucket2[k2] = it
                # 统计同文本出现次数（极端情况下限制同文本最多出现5次，
                # 即使位置不同也认为是叠印重复的同一文本行）
                cnt_key = it["text"]
                text_count[cnt_key] = text_count.get(cnt_key, 0) + 1
                if text_count[cnt_key] <= 5:
                    pass  # OK
            else:
                if it["prob"] > bucket2[k2]["prob"]:
                    bucket2[k2] = it
        deduplicated = list(bucket2.values())
    
    # 再次极端保护：如果还是>300个，对相同text的位置出现次数限制在3次以内
    if len(deduplicated) > 300:
        final_list = []
        text_counter = {}
        # 先按置信度从高到低排序，再截断同文本出现次数
        deduplicated.sort(key=lambda x: x["prob"], reverse=True)
        for it in deduplicated:
            cnt_key = it["text"]
            c = text_counter.get(cnt_key, 0)
            if c < 3:
                text_counter[cnt_key] = c + 1
                final_list.append(it)
        deduplicated = final_list
    
    # 给每个item添加分类（type/nums）
    all_items = []
    for it in deduplicated:
        item = dict(it)  # copy: text, x, y, w, h, prob, source
        item["type"] = "unknown"
        
        text_content = item["text"]
        lower = text_content
        
        # 1. 检测是否为关键字（红球/蓝球等变体）
        keyword_patterns = [
            ("red_dan", ["红胆", "红肛", "红旦"]),
            ("red_tuo", ["红拖", "红扡", "红施"]),
            ("red_ball", ["红球", "红爬", "红皿", "红昭", "红粑", "纤球", "红色", "醯球"]),
            ("red_ball_colon", ["红:", "红：", "红;", "红；"]),
            ("blue_ball", ["蓝球", "蓝胆", "蓝琛", "篮球", "兰球", "蓝求",
                           "蓝撇", "蓝搔", "蓝监", "蓝益", "蓝盔", "蓝盛", "蓝盟"]),
            ("blue_ball_colon", ["蓝:", "蓝：", "蓝;", "蓝；", "B球", "球:"]),
        ]
        is_keyword = False
        for kw_type, patterns in keyword_patterns:
            for p in patterns:
                if p in lower:
                    item["type"] = kw_type
                    is_keyword = True
                    break
            if is_keyword:
                break
        
        # 2. 检测倍数
        if not is_keyword:
            times_pats = [r"(\d+)倍", r"(\d+)倚", r"(\d+)份", r"(\d+)服", r"(\d+)股",
                          r"(\d+)们", r"(\d+)伯", r"(\d+)侪", r"(\d+)倌", r"(\d+)倩",
                          r"(\d+)宿", r"(\d+)佾", r"(\d+)伤", r"(\d+)府"]
            for tp in times_pats:
                m = re.search(tp, lower)
                if m:
                    try:
                        t = int(m.group(1))
                        if 1 <= t <= 100:
                            item["type"] = "times"
                            item["num"] = t
                            is_keyword = True
                            break
                    except:
                        pass
        
        # 2.5 噪音文本过滤：检测时间戳/电话号码/价格/装饰文字等非彩票号码
        if not is_keyword and _is_noise_ocr_text(text_content):
            # 噪音文本（时间戳、电话号、价格等）不提取号码
            all_items.append(item)
            continue

        # 3. 提取数字（两位数颠倒修正 + 连写拆分）
        if not is_keyword:
            num_matches = list(re.finditer(r"(\d+)", text_content))
            if num_matches:
                nums_extracted = []
                for nm in num_matches:
                    raw = nm.group(1)
                    # 使用连写拆分函数（在函数外部不可用，这里简化处理）
                    if len(raw) <= 2:
                        try:
                            n = int(raw)
                            if 10 <= n <= 99:
                                rev = int(str(n)[::-1])
                                if (1 <= rev <= 33) and not (1 <= n <= 33):
                                    n = rev
                            if 1 <= n <= 33:
                                nums_extracted.append(n)
                        except:
                            pass
                    else:
                        # 3-4位数字拆分（如0607→06,07）
                        L = len(raw)
                        if L == 3:
                            try:
                                a = int(raw[0]); b = int(raw[1:])
                                if 1 <= a <= 33 and 1 <= b <= 33:
                                    nums_extracted.extend([a, b]); continue
                            except: pass
                            try:
                                a = int(raw[:2]); b = int(raw[2])
                                if 1 <= a <= 33 and 1 <= b <= 33:
                                    nums_extracted.extend([a, b]); continue
                            except: pass
                        if L == 4:
                            try:
                                a = int(raw[:2]); b = int(raw[2:])
                                if 1 <= a <= 33 and 1 <= b <= 33:
                                    nums_extracted.extend([a, b]); continue
                            except: pass
                        # 每2位尝试拆分
                        i = 0
                        while i < L - 1:
                            try:
                                n = int(raw[i:i+2])
                                if 10 <= n <= 99:
                                    rev = int(str(n)[::-1])
                                    if (1 <= rev <= 33) and not (1 <= n <= 33):
                                        n = rev
                                if 1 <= n <= 33:
                                    nums_extracted.append(n)
                                    i += 2
                                    continue
                            except: pass
                            # 单数字
                            try:
                                n = int(raw[i])
                                if 1 <= n <= 33:
                                    nums_extracted.append(n)
                            except: pass
                            i += 1
                        if i == L - 1:
                            try:
                                n = int(raw[i])
                                if 1 <= n <= 33:
                                    nums_extracted.append(n)
                            except: pass
                if nums_extracted:
                    item["type"] = "number"
                    item["nums"] = nums_extracted
                    is_keyword = True
        
        all_items.append(item)
    
    # 按y排序，再按x排序，得到阅读顺序
    all_items.sort(key=lambda it: (round(it["y"]/20), it["x"]))
    
    # 生成纯文本行列表（用于传统解析器）- 过滤噪音文本
    lines_dict = defaultdict(list)
    for it in all_items:
        # 跳过噪音文本项（时间戳/电话号/价格等）
        if _is_noise_ocr_text(it.get("text", "")):
            continue
        y_bucket = round(it["y"]/30)  # 每行约30像素高度
        lines_dict[y_bucket].append(it)
    lines_text = []
    for yk in sorted(lines_dict.keys()):
        items_in_line = sorted(lines_dict[yk], key=lambda x: x["x"])
        line_text = " ".join([it["text"] for it in items_in_line if it["text"].strip()])
        if line_text.strip():
            lines_text.append(line_text.strip())
    
    return lines_text, all_items


# ===================== 3.1 组结果智能去重 + 质量过滤（多策略合并时用） =====================
def merge_unique_groups(group_lists, debug=False, keep_n=None):
    """合并多个策略的组结果，智能去重 + 质量过滤
    
    去重规则（收紧：相似组合并成1个，而不是都保留）：
    1. 红球集合完全相同 → 同一组，蓝球取并集
    2. 红球Jaccard>0.95 且 蓝球Jaccard>0.3 → 同一组（OCR变体）
    3. 红球Jaccard>0.85 且 蓝球Jaccard>0.7 → 同一组
    
    质量过滤（最终组数>keep_n时触发）：
    给每个组打分（含出现频次加分），高分组优先
    """
    # 先展平所有组，同时记录出现频次
    flat = []       # 组对象列表（去红球全同后）
    freq_map = {}   # red_tuple -> 出现的策略索引数（计数）
    for gi, gl in enumerate(group_lists):
        for g in gl:
            if not (6 <= len(g.get("red", [])) <= 30 and 1 <= len(g.get("blue", [])) <= 16):
                continue
            rk = tuple(sorted(g["red"]))
            if rk in freq_map:
                # 已有记录：收集蓝球集合（不直接union，避免噪音累积）
                exist = freq_map[rk]
                exist["_blue_sets"].append(set(g["blue"]))
                exist["times"] = max(exist.get("times", 1), g.get("times", 1))
                # 记录不同策略索引（用于频次）
                if gi not in exist["_strats"]:
                    exist["_strats"].add(gi)
            else:
                new_g = dict(g)
                new_g["_strats"] = {gi}  # 记录在哪些策略出现
                new_g["_blue_sets"] = [set(new_g["blue"])]  # 收集所有蓝球集合
                freq_map[rk] = new_g
    # 转成list
    flat = list(freq_map.values())

    if not flat:
        return []

    # ===== 选择最佳蓝球集合（避免union导致蓝球累积噪音）=====
    for g in flat:
        blue_sets = g.get("_blue_sets", [set(g["blue"])])
        if len(blue_sets) == 1:
            g["blue"] = sorted(list(blue_sets[0]))
        else:
            # 统计每个蓝球在多少个策略中出现
            blue_freq = {}
            for bs in blue_sets:
                for b in bs:
                    blue_freq[b] = blue_freq.get(b, 0) + 1
            # 优先选择1-3个蓝球的集合（复式蓝球通常1-3个）
            reasonable_sets = [bs for bs in blue_sets if 1 <= len(bs) <= 3]
            if reasonable_sets:
                # 在合理集合中，选频次总和最高的
                best_set = max(reasonable_sets, key=lambda bs: sum(blue_freq.get(b, 0) for b in bs))
                g["blue"] = sorted(list(best_set))
            else:
                # 所有集合都有4+蓝球（可能都是噪音）
                # 1. 尝试交集
                intersection = set.intersection(*blue_sets) if blue_sets else set()
                if 1 <= len(intersection) <= 4:
                    g["blue"] = sorted(list(intersection))
                else:
                    # 2. 选蓝球最少的集合，并截断到4个
                    smallest = min(blue_sets, key=len)
                    # 按频次排序取前4个
                    sorted_blues = sorted(smallest, key=lambda b: -blue_freq.get(b, 0))
                    g["blue"] = sorted(sorted_blues[:4])
        # 清理辅助字段
        if "_blue_sets" in g:
            del g["_blue_sets"]

    if debug:
        # 打印频次统计
        multi_freq = sum(1 for g in flat if len(g["_strats"]) >= 2)
        print(f"  [预去重-红球全同] {sum(len(s) for s in [g['_strats'] for g in flat])}次识别 → {len(flat)}组 (跨策略≥2次={multi_freq})")
    
    def jaccard(a, b):
        sa, sb = set(a), set(b)
        if not sa and not sb:
            return 1.0
        if not sa or not sb:
            return 0.0
        return len(sa & sb) / len(sa | sb)
    
    # 第二步：Jaccard聚类 - 高相似度合并（收紧阈值）
    clusters = []
    used = [False] * len(flat)
    for i in range(len(flat)):
        if used[i]:
            continue
        cluster = [i]
        used[i] = True
        for j in range(i + 1, len(flat)):
            if used[j]:
                continue
            red_sim = jaccard(flat[i]["red"], flat[j]["red"])
            blue_sim = jaccard(flat[i]["blue"], flat[j]["blue"])
            # 收紧的合并条件（比之前严格得多）
            same_group = False
            if red_sim >= 0.95 and blue_sim >= 0.3:
                same_group = True
            elif red_sim >= 0.85 and blue_sim >= 0.7:
                same_group = True
            if same_group:
                cluster.append(j)
                used[j] = True
        clusters.append(cluster)
    
    # 每个聚类：选最佳组（而非取并集，避免噪音累积）
    merged_list = []
    for cl in clusters:
        # 评估每个组的质量，选最佳的
        def _group_quality(idx):
            """组质量分：红球数量合理(6-8)+蓝球数量合理(1-3)+频次高"""
            g = flat[idx]
            s = 0
            rc, bc = len(g["red"]), len(g["blue"])
            # 红球数量分（6-8最佳）
            if 6 <= rc <= 8:
                s += 40
            elif rc == 5 or 9 <= rc <= 10:
                s += 25
            elif 11 <= rc <= 12:
                s += 10
            else:
                s += 5
            # 蓝球数量分（1-3最佳）
            if 1 <= bc <= 3:
                s += 30
            elif bc == 4:
                s += 20
            elif 5 <= bc <= 6:
                s += 10
            else:
                s += 3
            # 噪音模式惩罚：连续小数字（1,2,3,4）
            small_consec = sum(1 for n in [1,2,3,4,5] if n in g["red"])
            if small_consec >= 4:
                s -= 15
            elif small_consec >= 3:
                s -= 8
            # 跨策略频次加分
            f = len(g.get("_strats", set()))
            if f >= 3:
                s += 30
            elif f >= 2:
                s += 20
            else:
                s += 5
            return s

        # 选最佳组
        best_idx = max(cl, key=_group_quality)
        best_g = flat[best_idx]
        max_times = max(flat[idx].get("times", 1) for idx in cl)
        merged_strats = set()
        for idx in cl:
            merged_strats.update(flat[idx].get("_strats", set()))

        mr = sorted(set(best_g["red"]))
        # ===== 智能蓝球选择：从聚类中所有组收集蓝球集合，选最佳 =====
        all_blue_sets = [set(flat[idx]["blue"]) for idx in cl]
        if len(all_blue_sets) == 1:
            mb = sorted(list(all_blue_sets[0]))
        else:
            # 统计每个蓝球在多少个组中出现
            blue_freq = {}
            for bs in all_blue_sets:
                for b in bs:
                    blue_freq[b] = blue_freq.get(b, 0) + 1
            # 优先选择1-3个蓝球的集合
            reasonable_sets = [bs for bs in all_blue_sets if 1 <= len(bs) <= 3]
            if reasonable_sets:
                best_bs = max(reasonable_sets, key=lambda bs: sum(blue_freq.get(b, 0) for b in bs))
                mb = sorted(list(best_bs))
            else:
                # 交集
                intersection = set.intersection(*all_blue_sets) if all_blue_sets else set()
                if 1 <= len(intersection) <= 4:
                    mb = sorted(list(intersection))
                else:
                    # 选蓝球最少的，截断到4个
                    smallest = min(all_blue_sets, key=len)
                    sorted_blues = sorted(smallest, key=lambda b: -blue_freq.get(b, 0))
                    mb = sorted(sorted_blues[:4])
        overlap = set(mr) & set(mb)
        if overlap:
            mb = [b for b in mb if b not in overlap]
        if 6 <= len(mr) <= 30 and 1 <= len(mb) <= 16:
            merged_list.append({
                "red": mr,
                "blue": mb,
                "times": max_times,
                "_freq": len(merged_strats)  # 跨多少个不同策略识别到
            })
    
    if debug:
        print(f"  [去重合并] {len(flat)}组 → {len(merged_list)}组")
    
    # 第三步：质量过滤（当数量超过keep_n时）
    # keep_n自适应（关键：叠印严重=跨策略重复少，应该保守用更宽的keep_n）
    if keep_n is None:
        # 计算跨策略≥2次的高质量组数量
        robust_n = sum(1 for g in merged_list if g.get("_freq", 1) >= 2)
        # 各策略中产出过的最大单组数量
        max_single_strategy = max((len(gl) for gl in group_lists), default=0)
        
        if robust_n >= 8:
            # 跨策略重复多=关键字/空间信息清晰，最大单策略产出通常最接近实际
            # 所以用 max_single_strategy + 2~3 缓冲，不要用robust_n
            keep_n = max(8, max_single_strategy + 2)
            keep_n = min(keep_n, 13)
        elif robust_n >= 4:
            # 中等重复度：叠印较重，真实组可能跨策略重复少
            # 取 robust_n+5 和 最大策略+4 的较大值，至少保证12组（可能还有漏）
            keep_n = max(12, robust_n + 5, max_single_strategy + 4)
            keep_n = min(keep_n, 15)
        else:
            # robust_n<4: 叠印极端严重，真实组也可能只在1个策略里出现
            # 取"各策略最大产出"+5缓冲，并保留所有robust组
            keep_n = max(12, max_single_strategy + 5)
            keep_n = min(keep_n, 18)  # 叠印严重最多保留18组
    
    if len(merged_list) > keep_n:
        def _quality_score(g):
            """组质量分 + 出现频次（越高越好）"""
            s = 0
            rc = len(g["red"])
            bc = len(g["blue"])
            # 红球数量分（40）- 偏好6-8个红球
            if 6 <= rc <= 8:
                s += 40
            elif rc == 5 or 9 <= rc <= 10:
                s += 25
            elif 11 <= rc <= 12:
                s += 10
            else:
                s += 5
            # 蓝球数量分（30）- 偏好1-3个蓝球
            if 1 <= bc <= 3:
                s += 30
            elif bc == 4:
                s += 20
            elif 5 <= bc <= 6:
                s += 10
            else:
                s += 3
            # 跨度比（20）
            if rc > 0:
                rng = max(g["red"]) - min(g["red"])
                ratio = rng / max(rc, 1)
                if 1.5 <= ratio <= 3.0:
                    s += 20
                elif 1.0 <= ratio <= 4.0:
                    s += 15
                else:
                    s += 5
            else:
                s += 5
            # ===== 噪音模式惩罚（-15）=====
            # 连续小数字（1,2,3,4）通常是OCR噪音
            small_consecutive = 0
            for n in [1, 2, 3, 4, 5]:
                if n in g["red"]:
                    small_consecutive += 1
            if small_consecutive >= 4:
                s -= 15  # 4+个连续小数字=噪音
            elif small_consecutive >= 3:
                s -= 8
            # 跨策略频次加分（30）
            f = g.get("_freq", 1)
            if f >= 5:
                s += 30
            elif f >= 3:
                s += 25
            elif f >= 2:
                s += 18
            else:
                s += 5  # 只在1个策略出现
            return s
        
        merged_list.sort(key=lambda g: _quality_score(g), reverse=True)
        merged_list = merged_list[:keep_n]
        if debug:
            # 打印前5个的分
            top_scores = [_quality_score(g) for g in merged_list[:5]]
            print(f"  [质量过滤] >{keep_n}组→保留前{len(merged_list)}组 (top5分数={top_scores})")
    
    # 清理辅助字段
    for g in merged_list:
        if "_freq" in g:
            del g["_freq"]
    
    return merged_list


# ===================== 3.2 X方向分区预处理（解决左右分栏叠印） =====================
def split_items_x_zones(items, img_width=None, num_zones=None, debug=False):
    """将所有数字item按X方向分成多个区域（解决1896d7b3左中右三栏叠印）
    
    策略：
    1. 自动检测X方向的大间距（>80px）作为区分隔线
    2. 如果检测不到，则默认分2-3区（按图片宽度估计）
    """
    # 筛选有效数字（过滤低置信度噪音）
    numbers = []
    for it in items:
        if it["type"] == "number" and "nums" in it:
            # 过滤噪音文本项（时间戳/电话号/价格等）
            if "text" in it and _is_noise_ocr_text(it["text"]):
                continue
            # 过滤低置信度项（真实号码置信度通常>0.5，噪音<0.15）
            if it.get("prob", 0) < 0.15:
                continue
            for n in it["nums"]:
                if 1 <= n <= 33:
                    numbers.append({
                        "num": n,
                        "x": it["x"],
                        "y": it["y"],
                        "prob": it["prob"]
                    })
    if len(numbers) < 10:
        return [items]  # 太少，不分
    
    if not numbers:
        return [items]
    
    numbers.sort(key=lambda n: n["x"])
    
    # 方法1：检测X大间距
    x_gaps = []
    for i in range(1, len(numbers)):
        x_gaps.append(numbers[i]["x"] - numbers[i-1]["x"])
    
    if not x_gaps:
        return [items]
    
    avg_x_gap = sum(x_gaps) / len(x_gaps)
    gap_threshold = max(avg_x_gap * 3.5, 100)
    
    split_points = [0]
    for i, g in enumerate(x_gaps):
        if g > gap_threshold:
            split_points.append(numbers[i+1]["x"])
    split_points.append(float('inf'))
    
    if 2 <= len(split_points) - 1 <= 5:
        # 有明显的分区间距，按此分区
        if debug:
            print(f"  [X分区] 检测到 {len(split_points)-1} 个区域 (分割点X={split_points[1:-1]})")
    else:
        # 方法2：固定分区数（2-3区）
        min_x = numbers[0]["x"]
        max_x = numbers[-1]["x"]
        if img_width:
            max_x = max(max_x, img_width)
        if num_zones is None:
            num_zones = 3 if (max_x - min_x) > 500 else 2
        zone_width = (max_x - min_x) / num_zones
        split_points = [min_x + i * zone_width for i in range(num_zones)]
        split_points.append(float('inf'))
        if debug:
            print(f"  [X分区] 固定分 {num_zones} 区 (每区宽约{zone_width:.0f}px)")
    
    # 按split_points分items
    zones = [[] for _ in range(len(split_points) - 1)]
    for it in items:
        # 找出x对应的区域
        x = it.get("x", 0)
        for zi in range(len(split_points) - 1):
            if split_points[zi] <= x < split_points[zi + 1]:
                zones[zi].append(it)
                break
    
    # 过滤空区域
    zones = [z for z in zones if len(z) >= 3]
    # 所有item保底：确保不会因为分区导致空结果（加一个全量的兜底）
    zones.append(items)
    return zones


def _split_cluster_by_x(cluster_sorted, debug=False):
    """将一个Y方向聚类的号码按X方向间距拆分成多个子组

    当spatial_grouping的Y聚类跨多列时（如4×3网格的一行3列），
    需要按X间距把不同列的号码拆开。
    每个子组内部再按X间距分红球/蓝球。

    cluster_sorted: 已按x排序的数字列表 [{num, x, y, prob}, ...]
    返回: [(red_list, blue_list), ...]
    """
    if len(cluster_sorted) < 6:
        return []

    # 计算X间距
    x_gaps = []
    for i in range(1, len(cluster_sorted)):
        x_gaps.append(cluster_sorted[i]["x"] - cluster_sorted[i-1]["x"])

    if not x_gaps:
        return []

    # 找大间距分割点（间距>60px认为是不同列）
    # 使用动态阈值：平均间距的3倍，至少60px
    avg_gap = sum(x_gaps) / len(x_gaps)
    split_threshold = max(avg_gap * 2.5, 60)

    split_indices = []
    for i, g in enumerate(x_gaps):
        if g > split_threshold:
            split_indices.append(i + 1)

    if not split_indices:
        # 没有大的X间距，不能拆分
        return []

    # 按分割点把cluster分成多个子组
    sub_groups = []
    start = 0
    for si in split_indices:
        sub = cluster_sorted[start:si]
        if len(sub) >= 3:
            sub_groups.append(sub)
        start = si
    # 最后一段
    last_sub = cluster_sorted[start:]
    if len(last_sub) >= 3:
        sub_groups.append(last_sub)

    if len(sub_groups) < 2:
        return []

    # 对每个子组，按X间距分红球/蓝球
    result = []
    for sub in sub_groups:
        sub_sorted = sorted(sub, key=lambda n: n["x"])
        # 找最大X间距分红/蓝
        if len(sub_sorted) >= 4:
            sub_x_gaps = []
            for i in range(1, len(sub_sorted)):
                sub_x_gaps.append(sub_sorted[i]["x"] - sub_sorted[i-1]["x"])
            max_sub_gap = max(sub_x_gaps)
            max_sub_idx = sub_x_gaps.index(max_sub_gap)
            # 如果最大间距足够大且左边>=4个号，分割
            if max_sub_gap > 40 and max_sub_idx >= 3 and max_sub_idx < len(sub_sorted) - 1:
                red_part = sub_sorted[:max_sub_idx+1]
                blue_part = sub_sorted[max_sub_idx+1:]
            else:
                # 用蓝球范围(1-16)法
                split_idx = len(sub_sorted)
                for i in range(len(sub_sorted)-1, -1, -1):
                    if sub_sorted[i]["num"] > 16:
                        split_idx = i + 1
                        break
                red_part = sub_sorted[:split_idx]
                blue_part = sub_sorted[split_idx:] if split_idx < len(sub_sorted) else []
        else:
            red_part = sub_sorted
            blue_part = []

        reds = sorted(list(set([n["num"] for n in red_part if 1 <= n["num"] <= 33])))
        blues = sorted(list(set([n["num"] for n in blue_part if 1 <= n["num"] <= 16])))

        # 交叉污染过滤
        overlap = set(reds) & set(blues)
        if overlap:
            blues = [b for b in blues if b not in overlap]

        if reds and (len(reds) >= 4 or len(blues) >= 1):
            result.append((reds, blues))

    return result


# ===================== 3. 基于空间位置的号码分组（Fallback策略） =====================
def spatial_grouping(items, debug=False):
    """当传统关键字解析失败时，使用数字的空间位置进行分组
    
    原理：彩票图片中，每组号码的数字在y方向上聚集在一起（通常2-3行），
    组与组之间有较大的y间距。每组内部红球在前、蓝球在后，或红球在上、蓝球在下。
    
    items: 带位置的所有item列表
    返回: groups 列表 [{red, blue, times}]
    """
    # 筛选有效数字（过滤低置信度噪音）
    numbers = []
    for it in items:
        if it["type"] == "number" and "nums" in it:
            # 过滤噪音文本项（时间戳/电话号/价格等）
            if "text" in it and _is_noise_ocr_text(it["text"]):
                continue
            # 过滤低置信度项（真实号码置信度通常>0.5，噪音<0.15）
            if it.get("prob", 0) < 0.15:
                continue
            for n in it["nums"]:
                if 1 <= n <= 33:
                    numbers.append({
                        "num": n,
                        "x": it["x"],
                        "y": it["y"],
                        "prob": it["prob"]
                    })
    
    if len(numbers) < 8:  # 至少需要6红+1蓝+缓冲
        return []
    
    # 按y排序
    numbers.sort(key=lambda n: n["y"])
    
    # ===== Y方向聚类 =====
    # 计算相邻间距
    gaps = []
    for i in range(1, len(numbers)):
        gaps.append(numbers[i]["y"] - numbers[i-1]["y"])
    
    if not gaps:
        return []
    
    avg_gap = sum(gaps) / len(gaps)
    max_gap = max(gaps)
    
    # ===== 动态阈值自适应（叠印越严重阈值越低）=====
    # 总数字很多时（>80）说明叠印严重，降低阈值以分出更多组
    total_n = len(numbers)
    if total_n > 250:
        # 极端叠印（1896d7b3类型，叠印4+层）：阈值非常低，间距倍数1.2x
        y_threshold = max(avg_gap * 1.2, 18)
    elif total_n > 200:
        y_threshold = max(avg_gap * 1.4, 22)
    elif total_n > 150:
        # 非常严重叠印
        y_threshold = max(avg_gap * 1.6, 28)
    elif total_n > 100:
        # 严重叠印
        y_threshold = max(avg_gap * 1.8, 34)
    elif total_n > 60:
        # 中度叠印
        y_threshold = max(avg_gap * 2.1, 42)
    else:
        # 正常情况
        y_threshold = max(avg_gap * 2.6, 55)
    
    if debug:
        print(f"  [空间分组] 数字总数={total_n}, Y间距: avg={avg_gap:.1f}, max={max_gap:.1f}, 阈值={y_threshold:.1f}")
    
    y_clusters = []
    cur = [numbers[0]]
    for i in range(1, len(numbers)):
        gap = numbers[i]["y"] - numbers[i-1]["y"]
        if gap > y_threshold:
            y_clusters.append(cur)
            cur = [numbers[i]]
        else:
            cur.append(numbers[i])
    if cur:
        y_clusters.append(cur)
    
    # ===== 多级细分：大簇逐级用更细阈值分割 =====
    # 第1级细分割（阈值25-30px）：触发条件从>30降到>15
    def _refine_clusters(clusters, size_threshold, gap_threshold, level_name, dbg):
        refined = []
        for cl in clusters:
            if len(cl) > size_threshold:
                cl.sort(key=lambda n: n["y"])
                sub_cur = [cl[0]]
                for i in range(1, len(cl)):
                    gap = cl[i]["y"] - cl[i-1]["y"]
                    if gap > gap_threshold:
                        refined.append(sub_cur)
                        sub_cur = [cl[i]]
                    else:
                        sub_cur.append(cl[i])
                if sub_cur:
                    refined.append(sub_cur)
            else:
                refined.append(cl)
        if dbg and len(refined) != len(clusters):
            print(f"  [空间分组] {level_name}: {len(clusters)}→{len(refined)} (大小>{size_threshold}用阈值{gap_threshold}px)")
        return refined
    
    # 多级细化：叠印越严重，细分级数越多（触发更激进，阈值更细）
    if total_n > 250:
        # 极端叠印（1896d7b3）：5级细分，最小阈值到8px
        y_clusters = _refine_clusters(y_clusters, 15, 25, "二级细分", debug)
        y_clusters = _refine_clusters(y_clusters, 10, 18, "三级细分", debug)
        y_clusters = _refine_clusters(y_clusters, 7, 12, "四级细分", debug)
        y_clusters = _refine_clusters(y_clusters, 5, 9, "五级细分", debug)
    elif total_n > 150:
        # 非常严重叠印：4级
        y_clusters = _refine_clusters(y_clusters, 18, 28, "二级细分", debug)
        y_clusters = _refine_clusters(y_clusters, 12, 20, "三级细分", debug)
        y_clusters = _refine_clusters(y_clusters, 8, 14, "四级细分", debug)
    elif total_n > 60:
        # 中度/严重叠印：3级（原>80提前到>60）
        y_clusters = _refine_clusters(y_clusters, 20, 30, "二级细分", debug)
        y_clusters = _refine_clusters(y_clusters, 14, 22, "三级细分", debug)
    else:
        # 正常也保底做2级（>25用25）
        y_clusters = _refine_clusters(y_clusters, 25, 25, "三级细分", debug)
    
    if debug:
        print(f"  [空间分组] Y聚类数: {len(y_clusters)}")
        for ci, cl in enumerate(y_clusters):
            ys = [n["y"] for n in cl]
            xs = [n["x"] for n in cl]
            ns = sorted([n["num"] for n in cl])
            print(f"    聚类{ci+1}: {len(cl)}个数字 Y=[{min(ys):.0f},{max(ys):.0f}] X=[{min(xs):.0f},{max(xs):.0f}] nums={ns}")
    
    # ===== 对每个Y聚类，再按X方向分 红球区/蓝球区 =====
    groups = []
    for cluster in y_clusters:
        if len(cluster) < 3:  # 太少，跳过（可能是噪音）
            continue
        
        # 按x排序
        cluster_sorted = sorted(cluster, key=lambda n: n["x"])
        
        # 方法A: 找最大X间距作为红蓝球分隔
        if len(cluster_sorted) >= 2:
            x_gaps = []
            for i in range(1, len(cluster_sorted)):
                x_gaps.append(cluster_sorted[i]["x"] - cluster_sorted[i-1]["x"])
            
            max_x_gap = max(x_gaps)
            max_x_idx = x_gaps.index(max_x_gap)
            
            # 判断是否应该分割：最大间距 >= 其他间距均值的2倍，且分割后两边都有数字
            other_x_gaps = x_gaps[:max_x_idx] + x_gaps[max_x_idx+1:]
            avg_other_x = sum(other_x_gaps)/len(other_x_gaps) if other_x_gaps else 0
            
            should_split = (len(cluster_sorted) >= 8 and 
                          max_x_gap > 80 and 
                          max_x_idx >= 5 and  # 左边至少6个红球候选
                          max_x_idx < len(cluster_sorted) - 1)  # 右边至少1个蓝球
            
            if should_split and debug:
                print(f"    X间距: {[f'{g:.0f}' for g in x_gaps]}, max={max_x_gap:.0f}在idx={max_x_idx}, avg_other={avg_other_x:.0f}")
            
            if should_split:
                red_part = cluster_sorted[:max_x_idx+1]
                blue_part = cluster_sorted[max_x_idx+1:]
            else:
                # 方法B: 蓝球范围(1-16)过滤法 - 从右往左，连续的蓝球范围数字作为蓝球区
                # 从右往左找第一个不在蓝球范围的数字
                split_idx = len(cluster_sorted)
                for i in range(len(cluster_sorted)-1, -1, -1):
                    if cluster_sorted[i]["num"] > 16:
                        split_idx = i + 1
                        break
                if split_idx == len(cluster_sorted):
                    # 全部在蓝球范围，不分割
                    red_part = cluster_sorted
                    blue_part = []
                else:
                    red_part = cluster_sorted[:split_idx]
                    blue_part = cluster_sorted[split_idx:]
        else:
            red_part = cluster_sorted
            blue_part = []
        
        # 去重并验证红球(>=6个，1-33)和蓝球(>=1个，1-16)
        reds = sorted(list(set([n["num"] for n in red_part if 1 <= n["num"] <= 33])))
        blues = sorted(list(set([n["num"] for n in blue_part if 1 <= n["num"] <= 16])))
        
        # 如果红球太少，尝试放宽：把整个Y聚类都作为候选，按大小分
        if len(reds) < 6:
            all_nums_in_cluster = sorted(list(set([n["num"] for n in cluster_sorted if 1 <= n["num"] <= 33])))
            if len(all_nums_in_cluster) >= 7:
                # 先尝试用X间距最大的点，从后往前切蓝球
                for split_back in range(1, min(5, len(all_nums_in_cluster)-5)):
                    candidate_reds = sorted(all_nums_in_cluster[:-split_back])
                    candidate_blues = sorted([n for n in all_nums_in_cluster[-split_back:] if 1 <= n <= 16])
                    if len(candidate_reds) >= 6 and len(candidate_blues) >= 1:
                        # 交叉污染检查：红蓝球不应该有重叠
                        if len(set(candidate_reds) & set(candidate_blues)) == 0:
                            reds = candidate_reds
                            blues = candidate_blues
                            break

        if len(reds) >= 6 and len(blues) >= 1:
            # 交叉污染过滤
            overlap = set(reds) & set(blues)
            if overlap:
                # 有重叠，把重叠的从蓝球里移除（通常红球更多更可靠）
                blues = [b for b in blues if b not in overlap]
                if len(blues) < 1:
                    continue

            # ===== 红球过多时按X方向拆分（解决多列合并问题）=====
            # 真实复式组红球通常6-8个，>10说明可能是多列合并
            if len(reds) > 10 and len(cluster_sorted) >= 10:
                # 按X间距拆分成多个子组
                x_split_groups = _split_cluster_by_x(cluster_sorted, debug=debug)
                if x_split_groups:
                    for sub_reds, sub_blues in x_split_groups:
                        if 4 <= len(sub_reds) <= 10 and len(sub_blues) >= 0:
                            # 交叉污染检查
                            sub_overlap = set(sub_reds) & set(sub_blues)
                            if sub_overlap:
                                sub_blues = [b for b in sub_blues if b not in sub_overlap]
                            if len(sub_reds) >= 4:
                                groups.append({
                                    "red": sub_reds,
                                    "blue": sub_blues if sub_blues else [1],
                                    "times": 1
                                })
                                if debug:
                                    print(f"    → X拆分子组: 红球{sub_reds}, 蓝球{sub_blues}")
                    continue  # 跳过下面的整体添加

            # 数量合理性检查：红球6-15个，蓝球1-16个
            if 6 <= len(reds) <= 15 and 1 <= len(blues) <= 16:
                # 进一步检查：号码范围与数量的比例合理（避免多组合并过度）
                red_range = max(reds) - min(reds)
                red_count = len(reds)
                ratio = red_range / max(red_count, 1)
                range_ok = True
                if red_count >= 12 and ratio > 3.2:
                    range_ok = False
                elif red_count >= 8 and ratio > 4.5:
                    range_ok = False
                if not range_ok:
                    if debug:
                        print(f"    × 丢弃: 红球{min(reds)}-{max(reds)}跨度{red_range}数量{red_count} ratio={ratio:.1f}过大")
                    continue
                groups.append({
                    "red": reds,
                    "blue": blues,
                    "times": 1
                })
                if debug:
                    print(f"    → 有效组: 红球{reds}, 蓝球{blues}")
            else:
                if debug:
                    print(f"    × 跳过: 红球{len(reds)}个或蓝球{len(blues)}个不在合理范围")
    
    return groups


# ===================== 3.9 网格图片裁剪策略（多彩票拼图专用） =====================
def _detect_red_banner_regions(img):
    """检测图片顶部和底部的红色横幅区域，返回 (top_end, bottom_start)

    使用滑动窗口平均红色占比，避免文字区域导致检测中断。
    返回中间彩票区域的Y范围 [top_end, bottom_start]
    """
    h, w = img.shape[:2]
    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)

    # 检测红色像素（宽范围，覆盖暗红到亮红）
    lower_red1 = np.array([0, 40, 40])
    upper_red1 = np.array([20, 255, 255])
    lower_red2 = np.array([160, 40, 40])
    upper_red2 = np.array([180, 255, 255])
    red_mask = cv2.bitwise_or(
        cv2.inRange(hsv, lower_red1, upper_red1),
        cv2.inRange(hsv, lower_red2, upper_red2)
    )

    # 每行红色像素占比
    red_ratio = np.sum(red_mask > 0, axis=1) / w

    # 滑动窗口平均（窗口大小30）
    window = 30
    smoothed = np.convolve(red_ratio, np.ones(window)/window, mode='same')

    # 顶部横幅：从顶部开始，找到滑动平均降到0.25以下的位置
    top_end = 0
    for y in range(window, h - window):
        if smoothed[y] < 0.25:
            top_end = y
            break
    if top_end == 0:
        top_end = 0  # 没检测到横幅

    # 底部横幅：从底部开始，找到滑动平均降到0.25以下的位置
    bottom_start = h
    for y in range(h - window - 1, window, -1):
        if smoothed[y] < 0.25:
            bottom_start = y
            break
    if bottom_start == h:
        bottom_start = h

    # 确保中间区域足够大（至少占图片40%）
    if bottom_start - top_end < h * 0.4:
        top_end = 0
        bottom_start = h

    return int(top_end), int(bottom_start)


def _detect_grid_layout(img, top_offset=0, bottom_offset=None):
    """检测网格图片的行列数

    先去除顶部/底部红色横幅，再根据中间区域的宽高比和尺寸估算。
    """
    h, w = img.shape[:2]
    if bottom_offset is None:
        bottom_offset = h

    region_h = bottom_offset - top_offset
    region_w = w
    ratio = region_w / region_h if region_h > 0 else 1

    # 列数：大多数彩票拼图都是3列
    cols = 3

    # 行数：根据高度估算（每张彩票高度约200-240px）
    rows = max(2, round(region_h / 220))

    # 限制合理范围
    rows = min(6, max(2, rows))

    return rows, cols


def _split_consecutive_digits(num_str):
    """拆分连写的多位数字为1-33范围内的号码

    如 '1015' -> [10, 15], '912' -> [9, 12], '12426' -> [12, 26]
    """
    results = []
    s = num_str
    i = 0
    while i < len(s):
        # 尝试2位
        if i + 2 <= len(s):
            two = int(s[i:i+2])
            if 1 <= two <= 33:
                results.append(two)
                i += 2
                continue
        # 尝试1位
        one = int(s[i])
        if 1 <= one <= 9:
            results.append(one)
            i += 1
        else:
            i += 1
    return results


def _parse_sub_image_simple(sub_img, debug=False):
    """专门处理裁剪后的单张彩票子图：Y直方图峰值检测分红球/蓝球行

    策略：
    1. 收集所有数字（含连写拆分）
    2. Y方向直方图找两个密度峰值（红球行、蓝球行）
    3. 两峰之间最低点为分割线
    4. 频次统计提取号码
    """
    from collections import Counter

    h, w = sub_img.shape[:2]

    # 倍数标记替换模式
    times_sub_re = re.compile(r'\[?\d+\s*[倍份服股倚伯侪倌倩宿佾伤府佰]\]?')
    bracket_num_re = re.compile(r'\[\d+[^\]]*\]')

    # 收集所有数字：(num, y, prob)
    all_digits = []
    variants = preprocess_images(sub_img)
    for variant in variants:
        name, img = variant
        if name not in ("原图", "放大1.5x", "灰度图", "灰度+放大", "对比度增强", "二值化", "反色二值化", "自适应二值化"):
            continue
        try:
            results = reader.readtext(img, detail=1)
        except Exception:
            continue
        scale = 1.5 if "放大" in name else 1.0
        for bbox, text, prob in results:
            if not text or not text.strip():
                continue
            if _is_noise_ocr_text(text):
                continue
            if prob < 0.30:
                continue
            clean_text = times_sub_re.sub(' ', text)
            clean_text = bracket_num_re.sub(' ', clean_text)
            xs = [p[0] for p in bbox]
            ys = [p[1] for p in bbox]
            cy = sum(ys)/len(ys)/scale
            # 提取数字：先尝试连写拆分，再提取独立1-2位数字
            for match in re.finditer(r'\d+', clean_text):
                digit_str = match.group()
                if len(digit_str) <= 2:
                    try:
                        n = int(digit_str)
                        if 10 <= n <= 99:
                            rev = int(str(n)[::-1])
                            if (1 <= rev <= 33) and not (1 <= n <= 33):
                                n = rev
                        if 1 <= n <= 33:
                            all_digits.append((n, cy, prob))
                    except Exception:
                        pass
                else:
                    # 连写多位数字，拆分
                    for n in _split_consecutive_digits(digit_str):
                        all_digits.append((n, cy, prob))

    if len(all_digits) < 6:
        return None

    # ===== 鲁棒Y分割：去掉极值后在中间区域找最大间隔 =====
    ys_sorted = sorted([d[1] for d in all_digits])
    # 去掉最低和最高各15%的极值
    cut = int(len(ys_sorted) * 0.15)
    if cut > 0 and len(ys_sorted) > 2 * cut + 4:
        ys_trimmed = ys_sorted[cut:-cut]
    else:
        ys_trimmed = ys_sorted

    # 在修剪后的Y范围内找最大间隔
    max_gap = 0
    split_y = (ys_trimmed[0] + ys_trimmed[-1]) / 2
    for i in range(1, len(ys_trimmed)):
        gap = ys_trimmed[i] - ys_trimmed[i-1]
        if gap > max_gap:
            max_gap = gap
            split_y = (ys_trimmed[i] + ys_trimmed[i-1]) / 2

    # 校验：分割后上下两侧数字都不能太少
    red_count = sum(1 for d in all_digits if d[1] < split_y)
    blue_count = sum(1 for d in all_digits if d[1] >= split_y)
    if red_count < 3 or blue_count < 2:
        # 分割不合理，用中位数分割
        split_y = ys_sorted[len(ys_sorted)//2]

    # 上面红球，下面蓝球
    red_raw = [(n, y, p) for n, y, p in all_digits if y < split_y]
    blue_raw = [(n, y, p) for n, y, p in all_digits if y >= split_y]

    if not red_raw or not blue_raw:
        return None

    red_counter = Counter(n for n, _, _ in red_raw)
    blue_counter = Counter(n for n, _, _ in blue_raw if n <= 16)

    # 红球：优先取频次>=2的，限制6-8个
    red_by_freq = red_counter.most_common()
    red_high = [num for num, cnt in red_by_freq if cnt >= 2]
    if len(red_high) >= 5:
        red_selected = sorted(set(red_high[:8]))
    else:
        red_selected = sorted(set([num for num, _ in red_by_freq[:8]]))

    # 蓝球：优先取频次>=2的，限制1-4个
    blue_by_freq = [(num, cnt) for num, cnt in blue_counter.most_common() if num <= 16]
    blue_high = [num for num, cnt in blue_by_freq if cnt >= 2]
    if blue_high:
        blue_selected = sorted(set(blue_high[:4]))
    else:
        blue_selected = sorted(set([num for num, _ in blue_by_freq[:4]]))

    # 交叉污染过滤
    overlap = set(red_selected) & set(blue_selected)
    if overlap:
        blue_selected = [b for b in blue_selected if b not in overlap]

    # 降低阈值：红球>=1即可输出（避免子图因OCR质量差而完全失败）
    if len(red_selected) < 1 or len(blue_selected) < 1:
        if debug:
            print(f"    子图解析失败: split_y={split_y:.0f}, 红候选={red_counter.most_common(8)}, 蓝候选={blue_counter.most_common(5)}")
        return None

    if debug:
        print(f"    子图解析: split_y={split_y:.0f}")
        print(f"             红球候选={red_counter.most_common(10)}")
        print(f"             蓝球候选={blue_counter.most_common(5)}")
        print(f"             最终: 红={red_selected} 蓝={blue_selected}")

    return {"red": red_selected, "blue": blue_selected, "times": 1}


def _parse_grid_image(img, debug=False):
    """将网格图片裁剪成单张彩票子图，逐张识别后合并

    步骤：
    1. 检测并去除顶部/底部红色横幅
    2. 在中间彩票区域检测网格行列数
    3. 裁剪成子图，逐张用简单法解析，失败则回退
    """
    import tempfile
    h, w = img.shape[:2]

    # 1. 检测红色横幅，裁剪中间彩票区域
    top_end, bottom_start = _detect_red_banner_regions(img)
    if debug:
        print(f"  [网格裁剪] 红色横幅: 顶部[0,{top_end}], 底部[{bottom_start},{h}]")
    work_img = img[top_end:bottom_start, :]
    wh, ww = work_img.shape[:2]

    # 2. 检测网格布局
    rows, cols = _detect_grid_layout(work_img)
    if debug:
        print(f"  [网格裁剪] 中间区域 {ww}x{wh}, 检测到 {rows}行 x {cols}列 = {rows*cols} 张彩票")

    cell_w = ww / cols
    cell_h = wh / rows
    margin = 6  # 裁剪边距

    all_groups = []
    for r in range(rows):
        for c in range(cols):
            x1 = max(0, int(c * cell_w) + margin)
            y1 = max(0, int(r * cell_h) + margin)
            x2 = min(ww, int((c + 1) * cell_w) - margin)
            y2 = min(wh, int((r + 1) * cell_h) - margin)
            sub_img = work_img[y1:y2, x1:x2]
            if sub_img.size == 0 or sub_img.shape[0] < 50 or sub_img.shape[1] < 50:
                continue
            # 先用简单频次统计法解析
            g = _parse_sub_image_simple(sub_img, debug=debug)
            if g:
                all_groups.append(g)
                if debug:
                    print(f"    子图[{r},{c}] 简单法: 红={g['red']} 蓝={g['blue']}")
            else:
                # 失败回退到原有多策略逻辑
                tmp_path = os.path.join(tempfile.gettempdir(), f"ssq_sub_fb_{r}_{c}.png")
                cv2.imwrite(tmp_path, sub_img)
                fallback = parse_lottery_image(tmp_path, debug=False, _is_sub=True)
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass
                if fallback:
                    best = max(fallback, key=lambda x: len(x.get("red", [])))
                    all_groups.append(best)
                    if debug:
                        print(f"    子图[{r},{c}] 回退法: 红={best['red']} 蓝={best['blue']}")
                else:
                    if debug:
                        print(f"    子图[{r},{c}] 识别失败")
    return all_groups


# ===================== 4. 单图提取所有复式红球蓝球分组 =====================
def parse_lottery_image(img_path, debug=False, _is_sub=False):
    # 加载图片
    img = load_image(img_path)
    if img is None:
        return []

    h, w = img.shape[:2]
    if debug:
        print(f"  图片尺寸: {w}x{h}")

    # ===== 网格图片裁剪策略（仅顶层调用时执行，子图不再裁剪）=====
    # 当图片是多张彩票拼成的网格图（如4行3列共12张）时，
    # 先裁剪成单张彩票子图再逐张识别，避免号码互相干扰
    if not _is_sub and h >= 600 and w >= 800:
        grid_groups = _parse_grid_image(img, debug=debug)
        if grid_groups and len(grid_groups) >= 6:
            if debug:
                print(f"  [网格裁剪] 裁剪后共识别到 {len(grid_groups)} 组，采用网格策略结果")
            return grid_groups

    # 生成多种预处理版本
    img_variants = preprocess_images(img)
    
    # 对每种预处理执行OCR
    all_ocr_results = []
    for variant in img_variants:
        name, results = ocr_with_detail(variant)
        if results:
            all_ocr_results.append((name, results))
            if debug:
                print(f"  [{name}] 识别到 {len(results)} 个区块")
    
    if not all_ocr_results:
        return []
    
    # 合并所有OCR结果
    merged_lines, merged_items = merge_ocr_results(all_ocr_results)
    
    if debug:
        print(f"  合并后 {len(merged_lines)} 行文本, {len(merged_items)} 个item")
        if merged_lines:
            for i, l in enumerate(merged_lines[:50]):
                print(f"    [{i:3d}] {l}")
    
    # ===== 新策略：全并行运行 + 结果合并去重 =====
    # 所有策略都运行，组结果全部收集后合并去重（解决1896d7b3关键字4组+空间分组8组=12组的问题）
    all_strategy_groups = []
    
    # --- 策略A: 关键字解析 ---
    # A1: 合并文本解析
    groups_merged_kw = try_parse_lottery_lines(merged_lines, debug and False)
    all_strategy_groups.append(groups_merged_kw)
    if debug:
        print(f"  [A1:合并关键字] 解析到 {len(groups_merged_kw)} 组")
    
    # A2: 逐个预处理文本的关键字解析（每个预处理单独跑）
    for src_name, ocr_res in all_ocr_results:
        # 过滤噪音文本行（时间戳/电话号/价格等）
        lines_only = [text.strip() for _, text, _ in ocr_res 
                      if text.strip() and not _is_noise_ocr_text(text)]
        g = try_parse_lottery_lines(lines_only, debug and False)
        if g:
            all_strategy_groups.append(g)
            if debug and len(g) > 0:
                print(f"  [A2:{src_name}-关键字] 解析到 {len(g)} 组")
    
    # --- 策略B: 合并items的空间分组 ---
    spatial_merged = spatial_grouping(merged_items, debug=debug and False)
    all_strategy_groups.append(spatial_merged)
    if debug:
        print(f"  [B1:合并空间] 解析到 {len(spatial_merged)} 组")
    
    # --- 策略C: X分区 + 分区内空间分组（解决左右三栏叠印）---
    x_zones = split_items_x_zones(merged_items, img_width=w, debug=debug and False)
    if debug:
        print(f"  [X分区] 得到 {len(x_zones)} 个X区域（含全量兜底）")
    for zi, zone_items in enumerate(x_zones):
        sg_zone = spatial_grouping(zone_items, debug=False)
        if sg_zone:
            all_strategy_groups.append(sg_zone)
            if debug:
                print(f"  [C{zi+1}:X分区{zi+1}-空间] 解析到 {len(sg_zone)} 组")
    
    # --- 策略D: 逐个预处理单独空间分组（叠印少的预处理效果更好）---
    # 只跑主要预处理（原图/灰度/对比度增强/放大1.5x/灰度放大），二值化类容易出噪音暂不跑
    GOOD_PREPROCESS = {"原图", "灰度图", "对比度增强", "放大1.5x", "灰度+放大"}
    for src_name, ocr_res in all_ocr_results:
        if src_name not in GOOD_PREPROCESS:
            continue
        # 从该预处理结果重建items
        temp_items = []
        for bbox, text, prob in ocr_res:
            if not text.strip():
                continue
            # 过滤噪音文本（时间戳/电话号/价格等）
            if _is_noise_ocr_text(text):
                continue
            # 过滤低置信度项
            if prob < 0.15:
                continue
            xs = [p[0] for p in bbox]
            ys = [p[1] for p in bbox]
            cx, cy = sum(xs)/len(xs), sum(ys)/len(ys)
            nums_found = []
            for nm in re.finditer(r"(\d{1,2})", text):
                try:
                    n = int(nm.group(1))
                    if 10 <= n <= 99:
                        rev = int(str(n)[::-1])
                        if (1 <= rev <= 33) and not (1 <= n <= 33):
                            n = rev
                    if 1 <= n <= 33:
                        nums_found.append(n)
                except:
                    pass
            if nums_found:
                for n in nums_found:
                    temp_items.append({
                        "type": "number",
                        "nums": [n],
                        "num": n,
                        "x": cx, "y": cy, "prob": prob
                    })
        # D1: 该预处理整体空间分组
        sg_d = spatial_grouping(temp_items, debug=False)
        if sg_d:
            all_strategy_groups.append(sg_d)
            if debug:
                print(f"  [D1:{src_name}-空间] 解析到 {len(sg_d)} 组")
        # D2: 该预处理X分区空间分组（只在整体分组<8时启用，避免大图噪音爆炸）
        if len(sg_d) < 8:
            zones_d = split_items_x_zones(temp_items, img_width=w)
            for zone_items_d in zones_d[:-1]:  # 去掉最后一个全量兜底（和D1重复）
                sg_dz = spatial_grouping(zone_items_d, debug=False)
                if sg_dz:
                    all_strategy_groups.append(sg_dz)
                    if debug:
                        print(f"  [D2:{src_name}-X分区空间] 解析到 {len(sg_dz)} 组")
    
    # --- 合并所有策略结果并去重 + 质量过滤 ---
    total_flat = sum(len(gl) for gl in all_strategy_groups)
    if debug:
        print(f"  [汇总] 所有策略共产出 {total_flat} 组候选")

    final_groups = merge_unique_groups(all_strategy_groups, debug=debug)

    # ===== 最终修复：修复明显有噪音的组（而非直接删除，保留12组结构）=====
    def _repair_noise_group(g):
        """修复噪音组：截断过多蓝球、移除连续小数字红球"""
        reds = list(g["red"])
        blues = list(g["blue"])
        # 1. 蓝球过多（>4）→ 截断到前4个（复式蓝球通常1-4个）
        if len(blues) > 4:
            blues = blues[:4]
        # 2. 连续小数字红球（1,2,3,4,5中有4+个）→ 移除这些小数字（噪音）
        small_nums = [n for n in [1,2,3,4,5] if n in reds]
        if len(small_nums) >= 4:
            reds = [r for r in reds if r not in [1,2,3,4,5]]
        # 3. 红球过多（>12）→ 截断到12个（按大小排序保留前12）
        if len(reds) > 12:
            reds = sorted(reds)[:12]
        # 4. 交叉污染过滤
        overlap = set(reds) & set(blues)
        if overlap:
            blues = [b for b in blues if b not in overlap]
        # 5. 修复后红球不足6个 → 无效，返回None
        if len(reds) < 6:
            return None
        if len(blues) < 1:
            blues = [1]  # 兜底
        g["red"] = sorted(reds)
        g["blue"] = sorted(blues)
        return g

    # 对所有组进行修复（不删除，而是修复号码）
    repaired = []
    removed_count = 0
    for g in final_groups:
        repaired_g = _repair_noise_group(g)
        if repaired_g is not None:
            repaired.append(repaired_g)
        else:
            removed_count += 1
            # 修复后无效的组保留原样（避免组数减少）
            repaired.append(g)
    if debug and removed_count > 0:
        print(f"  [噪音修复] 修复{len(final_groups)}组, 其中{removed_count}组修复后仍无效（保留原样）")
    final_groups = repaired

    if debug:
        print(f"  [最终] 合并去重后 {len(final_groups)} 组")
        for gi, g in enumerate(final_groups):
            print(f"    组{gi+1}: 红{g['red']} ({len(g['red'])}个), 蓝{g['blue']} ({len(g['blue'])}个), 倍{g['times']}")

    return final_groups

def try_parse_lottery_lines(lines, debug=False):
    """尝试从OCR文本中解析双色球号码
    
    处理多种格式：
    1. 连续行格式: 红球:01 07 13 20 24 28 32 33
                  蓝球:10 13 14 [1倍]
    2. 分行格式:   红球:01
                  07
                  13
                  ...
                  蓝球:10
                  13
                  14
                  [1倍]
    3. 胆拖格式:   红胆:07 12 23 24
                  红拖:02 04 06 08 11 15 19
                  蓝球:01 02
                  [1倍]
    4. OCR误读格式: 红爬:02, 红皿:01, 蓝琛:09, [5倚] 等
    """
    
    # ====== 辅助函数（模块级，避免闭包变量问题） ======
    # 蓝球/红球提取时的截断关键词（遇到这些词立即停止，避免广告文字混入数字）
    # 注意：只把明确出现在号码区结束之后的词放进来！
    # 出现在号码中间的干扰词（如"心佑"、"魔K]"、"今晚"）不要放（它们后面还有号码）
    BLUE_STOP_WORDS = [
        "红球", "红胆", "红拖", "红爬", "纤球",  # 明确的红球关键字=蓝球区结束
        "双色", "双色球",  # 下一组彩票开始标记
        "[", "倍",  # 倍数标记开始=蓝球区结束（[1倍]中的数字不被提取）
        "C1倍", "C1倚",  # OCR误读的倍数标记
        # 以下一般出现在所有号码之后的广告/感谢区域
        "公益", "感谢您", "感谢你", "永州福地", "永州福",
        "必中款", "必_中", "暴富", "财源滚滚", "上岸双色球",
        "团队研究", "团队", "大奖", "祝福", "福地",
        "约定打卡", "打卡",
    ]
    RED_STOP_WORDS = [
        "蓝球", "蓝胆", "篮球", "兰球", "蓝求",  # 明确的蓝球关键字=红球区结束
        "双色", "双色球",  # 下一组彩票开始标记
        "[", "倍",  # 倍数标记开始=红球区结束
        "C1倍", "C1倚",
        "公益", "感谢您", "感谢你", "永州福地", "永州福",
        "必中款", "必_中", "暴富", "财源滚滚", "上岸双色球",
    ]
    
    def _fix_num_ocr(n):
        """OCR两位数颠倒修正（高频错误模式：92→29, 61→19, 81→18等）"""
        if 10 <= n <= 99:
            s = str(n)
            rev = int(s[::-1])
            # 常见OCR颠倒：原数字不在范围但颠倒后在范围 → 修正
            if not (1 <= n <= 33) and (1 <= rev <= 33):
                return rev
        return n
    
    def _split_joined_digits(raw_str, max_range=33):
        """处理OCR数字连写问题（如'0607'→06,07；'1415'→14,15；'2931'→29,31）
        
        策略：
        - 1-2位：正常处理
        - 3位：优先拆1位+2位（后2位在范围内则拆，否则前2位+1位）
        - 4位：优先2位+2位（都在范围内则拆）
        - 5-6位：递归2位切（每2位在范围内）
        """
        results = []
        s = raw_str.strip()
        if not s.isdigit():
            return results
        L = len(s)
        if L <= 2:
            n = _fix_num_ocr(int(s))
            if 1 <= n <= max_range:
                results.append(n)
            return results
        if L == 3:
            # 尝试 1+2 拆分
            a = _fix_num_ocr(int(s[0]))
            b = _fix_num_ocr(int(s[1:]))
            if 1 <= a <= max_range and 1 <= b <= max_range:
                results.extend([a, b])
                return results
            # 尝试 2+1
            a = _fix_num_ocr(int(s[:2]))
            b = _fix_num_ocr(int(s[2]))
            if 1 <= a <= max_range and 1 <= b <= max_range:
                results.extend([a, b])
                return results
            # 都不行，整体处理
            n = _fix_num_ocr(int(s))
            if 1 <= n <= max_range:
                results.append(n)
            return results
        if L == 4:
            a = _fix_num_ocr(int(s[:2]))
            b = _fix_num_ocr(int(s[2:]))
            if 1 <= a <= max_range and 1 <= b <= max_range:
                results.extend([a, b])
                return results
            # 尝试 2+1+1 等递归
            sub1 = _split_joined_digits(s[:2], max_range)
            sub2 = _split_joined_digits(s[2:], max_range)
            if sub1 and sub2:
                results.extend(sub1)
                results.extend(sub2)
                return results
            n = _fix_num_ocr(int(s))
            if 1 <= n <= max_range:
                results.append(n)
            return results
        # L>=5，每2位切
        i = 0
        failed = False
        while i < L:
            seg = s[i:i+2]
            if len(seg) == 1:
                n = _fix_num_ocr(int(seg))
                if 1 <= n <= max_range:
                    results.append(n)
                else:
                    failed = True
                i += 1
            else:
                n = _fix_num_ocr(int(seg))
                if 1 <= n <= max_range:
                    results.append(n)
                    i += 2
                else:
                    # 尝试1位
                    n1 = _fix_num_ocr(int(seg[0]))
                    if 1 <= n1 <= max_range:
                        results.append(n1)
                        i += 1
                    else:
                        failed = True
                        break
        if not failed:
            return results
        return []
    
    def extract_red_nums(text):
        """从文本中提取红球号码(1-33)，含OCR误读修正+连写拆分+关键词截断"""
        nums = []
        # 先截断：遇到红球停止词（如蓝球、公益等），之前的内容提取号码
        stop_idx = len(text)
        lower_t = text
        for sw in RED_STOP_WORDS:
            idx = lower_t.find(sw)
            if 0 <= idx < stop_idx:
                stop_idx = idx
        truncated = text[:stop_idx] if stop_idx < len(text) else text
        
        for m in re.finditer(r"(\d+)", truncated):
            raw = m.group(1)
            parsed = _split_joined_digits(raw, 33)
            for n in parsed:
                if 1 <= n <= 33 and n not in nums:
                    nums.append(n)
        return nums
    
    def extract_blue_nums(text):
        """从文本中提取蓝球号码(1-16)，含OCR误读修正+连写拆分+关键词截断"""
        nums = []
        # 先截断：遇到蓝球停止词（如红球、公益、感谢等），立即停止提取后续号码
        stop_idx = len(text)
        lower_t = text
        for sw in BLUE_STOP_WORDS:
            idx = lower_t.find(sw)
            if 0 <= idx < stop_idx:
                stop_idx = idx
        truncated = text[:stop_idx] if stop_idx < len(text) else text
        
        for m in re.finditer(r"(\d+)", truncated):
            raw = m.group(1)
            parsed = _split_joined_digits(raw, 16)
            for n in parsed:
                if 1 <= n <= 16 and n not in nums:
                    nums.append(n)
        return nums
    
    def extract_times(text):
        """从文本中提取倍数（份/服/股/倍 等OCR变体）"""
        # 扩展更多倍数关键词：服、份、股（图片中常混淆）
        for pat in [
            r"\[?(\d+)倍\]?", r"\[?(\d+)倚\]?", r"(\d+)倍", r"(\d+)倚",
            r"\[?(\d+)服\]?", r"(\d+)服",
            r"\[?(\d+)份\]?", r"(\d+)份",
            r"\[?(\d+)股\]?", r"(\d+)股",
            r"\[?(\d+)们\]?", r"\[?(\d+)伯\]?", r"\[?(\d+)侪\]?",
            r"\[?(\d+)倌\]?", r"\[?(\d+)倩\]?", r"\[?(\d+)宿\]?",
            r"\[?(\d+)佾\]?", r"\[?(\d+)伤\]?", r"\[?(\d+)府\]?",
        ]:
            m = re.search(pat, text)
            if m:
                try:
                    t = int(m.group(1))
                    if 1 <= t <= 100:
                        return t
                except:
                    pass
        return None
    
    # ====== 状态变量 ======
    all_groups = []
    
    reading_red = [False]      # 是否正在读取红球
    reading_blue = [False]     # 是否正在读取蓝球
    current_red = [[]]         # 当前红球列表
    current_blue = [[]]        # 当前蓝球列表
    current_times = [1]        # 当前倍数
    current_red_dan = [[]]     # 胆拖格式-胆码
    current_red_tuo = [[]]     # 胆拖格式-拖码
    in_dantuo_mode = [False]   # 是否胆拖模式
    passed_tuo = [False]       # 胆拖模式下是否已过"红拖"标记
    # 红球队列：当多个红球组连续出现时，先缓存红球队列，等待蓝球配对
    pending_red_groups = []    # 每个元素: {"red": [...], "times": N, "dantuo": bool, "dan": [...], "tuo": [...]}
    # ===== 广告区过滤标志 =====
    # 倍数行出现 → 号码区结束 → 进入广告区。之后只有遇到显式红球/蓝球关键字才能重新进入号码区。
    # 防止广告行（"喜中X期"、"X注"、"奖池X"、历史号码等）的数字被识别成彩票号码。
    in_ad_zone = [False]
    
    def _is_red_keyword(line):
        """检测行是否以红球/红胆/红拖等关键字开头，返回类型和同行数字
        
        注意：必须从最具体的关键字开始匹配，避免"红"先匹配了"红胆/红球/红拖"等
        """
        # 红球关键字变体（从最具体到最不具体排序）
        # 新增更多OCR误读：纤球(=红球)、红旦(=红胆)、红施(=红拖)、红肛(=红胆)
        # 支持分号;作为冒号:的误读（红球;05 = 红球:05）
        red_keys = [
            ("红球", False),   # 必须先于"红"
            ("红胆", True),    # 胆码
            ("红拖", False),   # 拖码
            ("红爬", False),   # OCR误读
            ("红皿", False),   # OCR误读
            ("红昭", False),   # OCR误读
            ("红粑", False),   # OCR误读
            ("纤球", False),   # OCR误读 (纟+工=红的误读)
            ("红旦", True),    # OCR误读 (胆→旦)
            ("红扡", False),   # OCR误读 (拖→扡)
            ("红施", False),   # OCR误读 (拖→施)
            ("红肛", True),    # OCR误读 (胆→肛)
            ("红色", False),   # 兜底：红色=红球
            ("醯球", False),   # OCR误读 (红球→醯球)
            ("红", False),     # 兜底：单字红（必须带冒号以避免误匹配）
        ]
        for keyword, is_dan in red_keys:
            # 对每个关键字，尝试 关键字+冒号/分号/空格，或直接从行开头匹配
            # 支持 : ： ; ； 四种分隔符（OCR常把冒号识别成分号）
            # 红球使用短匹配（连续数字+空格）：避免叠印把后面蓝球区/广告数字混入
            # 后面如果还有红球号码会通过延续行extra追加补充
            if keyword == "红":
                # 单字红必须带冒号/分号，避免匹配"红胆"、"红球"等
                pat = re.compile(r"红[:：;；]\s*((?:\d{1,2}\s*)*)")
            else:
                # 允许 : ： ; ； 或 空格 直接跟数字
                pat = re.compile(re.escape(keyword) + r"[:：;；]?\s*((?:\d{1,2}\s*)*)")
            m = pat.match(line)
            if m:
                nums_text = m.group(1) if m.lastindex and m.lastindex >= 1 else ""
                nums = extract_red_nums(nums_text)
                return is_dan, nums
        # 兜底：行内包含"红球/红胆"等关键字但不是从开头（前面有空格/特殊字符）
        for keyword, is_dan in [("红球", False), ("红胆", True), ("红拖", False), 
                                 ("纤球", False), ("红旦", True), ("红肛", True),
                                 ("醯球", False)]:
            idx = line.find(keyword)
            if idx >= 0:
                rest = line[idx+len(keyword):]
                # 去掉前面的冒号/分号/空格
                rest = re.sub(r"^[:：;；\s]+", "", rest)
                nums = extract_red_nums(rest)
                return is_dan, nums
        return None, []
    
    def _is_blue_keyword(line):
        """检测行是否以蓝球关键字开头，返回同行数字
        
        注意：必须从最具体的关键字开始匹配
        """
        # 新增更多OCR误读：篮球(=蓝球)、兰球(=蓝球)、蓝求(=蓝球)
        # 蓝撇=蓝球、蓝l0l=蓝球:01（OCR把球:01识别成l0l）、蓝IBI=蓝球:01
        # 蓝面、蓝监、蓝益 等（"球"字的各种OCR误读）
        blue_keys = [
            "蓝球", "蓝胆", "蓝琛", "篮球", "兰球", "蓝求", "篮胆", "兰胆",
            "蓝撇", "蓝搔", "蓝监", "蓝益", "蓝盔", "蓝盛", "蓝盟",
            "蓝l0l", "蓝101", "蓝IBI", "蓝ibi", "蓝L0L", "蓝ioi",
            "蓝l01", "蓝10l", "蓝lol", "蓝L0l",
            "蓝",
        ]
        for keyword in blue_keys:
            if keyword == "蓝":
                # 单字蓝必须带冒号/分号
                pat = re.compile(r"蓝[:：;；]\s*(.*)$")
            else:
                # 支持 : ： ; ； 四种分隔符
                # ===== 关键修复：关键字后匹配整行，遇到心佑/匐;-等干扰词不会提前停止 =====
                pat = re.compile(re.escape(keyword) + r"[:：;；]?\s*(.*)$")
            m = pat.match(line)
            if m:
                nums_text = m.group(1) if m.lastindex and m.lastindex >= 1 else ""
                nums = extract_blue_nums(nums_text)
                # 特殊处理：蓝l0l / 蓝101 这种情况，关键字本身的"l0l"是"球:01"的误读
                # 如果匹配到但没有数字，尝试从关键字的后几位提取
                if not nums and len(keyword) > 2:
                    tail = keyword[2:]  # 去掉"蓝"字后的部分
                    for nm in re.finditer(r"(\d{1,2})", tail):
                        try:
                            n = _fix_num_ocr(int(nm.group(1)))
                            if 1 <= n <= 16:
                                nums.append(n)
                        except:
                            pass
                if nums:
                    return nums
        # 兜底：匹配 "B球" 或 "球:" 开头（缺少"蓝"字的情况）
        for pat_str in [r"[Bb]球[:：;；]?\s*((?:\d{1,2}\s*)*)",
                        r"球[:：;；]\s*((?:\d{1,2}\s*)*)"]:
            m = re.match(pat_str, line)
            if m:
                nums_text = m.group(1) if m.lastindex and m.lastindex >= 1 else ""
                nums = extract_blue_nums(nums_text)
                return nums
        # 兜底：行内出现关键字（不限于开头）
        for keyword in ["蓝球", "蓝胆", "篮球", "兰球", "蓝求", "蓝撇", "蓝监", "蓝益",
                        "蓝琛", "蓝l0l", "蓝101", "蓝lol"]:
            idx = line.find(keyword)
            if idx >= 0:
                rest = line[idx+len(keyword):]
                rest = re.sub(r"^[:：;；\s]+", "", rest)
                nums = extract_blue_nums(rest)
                # 同样处理关键字尾巴带数字的情况
                if not nums and len(keyword) > 2:
                    tail = keyword[2:]
                    for nm in re.finditer(r"(\d{1,2})", tail):
                        try:
                            n = _fix_num_ocr(int(nm.group(1)))
                            if 1 <= n <= 16 and n not in nums:
                                nums.append(n)
                        except:
                            pass
                if nums:
                    return nums
        return []
    
    def _save_current_red():
        """将当前已读取的红球保存到等待队列（等待蓝球配对）
        阈值降低到>=1个号码：因为关键字行可能只有1个号码，后续延续行会继续追加
        """
        if in_dantuo_mode[0]:
            total = len(current_red_dan[0]) + len(current_red_tuo[0])
            if total >= 1:
                pending_red_groups.append({
                    "dantuo": True,
                    "dan": list(current_red_dan[0]),
                    "tuo": list(current_red_tuo[0]),
                    "times": current_times[0]
                })
                # 重置红球相关状态
                current_red_dan[0] = []
                current_red_tuo[0] = []
                in_dantuo_mode[0] = False
                passed_tuo[0] = False
                current_times[0] = 1
                reading_red[0] = False
        else:
            if len(current_red[0]) >= 1:
                pending_red_groups.append({
                    "dantuo": False,
                    "red": list(current_red[0]),
                    "times": current_times[0]
                })
                current_red[0] = []
                current_times[0] = 1
                reading_red[0] = False
    
    def _append_to_last_pending_or_current(nums, is_red, is_blue=False):
        """当有号码延续行时，追加到：
        1. 如果正在reading_red/reading_blue → 追加到current_red/current_blue
        2. 否则如果pending队列最后一组还没配对 → 追加到pending最后一组的red
        这样避免：遇到新红球关键字_save_current_red后，后续延续号码无处追加
        
        【防叠印限制】：
        - 普通红球模式（非胆拖）：追加后总数>10时，拒绝额外号码（1-33中真正的复式组一般6-10个红球封顶，>10基本是叠印污染）
        - 胆拖模式：不限制（胆2拖N的N可以很大）
        - 蓝球：>16拒绝（最大16个）
        """
        if not nums:
            return
        if is_red:
            if reading_red[0] and len(current_red[0]) > 0:
                # 普通模式+已经有号码：先算追加后总数，超限就跳过
                if not in_dantuo_mode[0]:
                    cap = 10
                    remaining = cap - len(current_red[0])
                    if remaining <= 0:
                        return
                    # 按原顺序取remaining个
                    seen = set(current_red[0])
                    to_add = []
                    for n in nums:
                        if len(to_add) >= remaining:
                            break
                        if 1 <= n <= 33 and n not in seen:
                            to_add.append(n); seen.add(n)
                    if to_add:
                        _add_red_nums(to_add)
                    return
                _add_red_nums(nums)
                return
            # 追加到pending队列的最后一组
            if pending_red_groups:
                last = pending_red_groups[-1]
                if last["dantuo"]:
                    for n in nums:
                        if n not in last["tuo"] and n not in last["dan"] and 1 <= n <= 33:
                            last["tuo"].append(n)
                else:
                    cap = 10
                    cur_len = len(last["red"])
                    if cur_len >= cap:
                        return
                    remaining = cap - cur_len
                    seen = set(last["red"])
                    added = 0
                    for n in nums:
                        if added >= remaining:
                            break
                        if n not in seen and 1 <= n <= 33:
                            last["red"].append(n); seen.add(n); added += 1
                return
            # 都没有，暂存到current_red
            if not reading_red[0]:
                reading_red[0] = True
            if not in_dantuo_mode[0]:
                cap = 10
                seen = set(current_red[0])
                to_add = []
                for n in nums:
                    if len(to_add) + len(current_red[0]) >= cap:
                        break
                    if 1 <= n <= 33 and n not in seen:
                        to_add.append(n); seen.add(n)
                _add_red_nums(to_add)
            else:
                _add_red_nums(nums)
        else:  # blue
            cap = 16
            if len(current_blue[0]) >= cap:
                return
            remaining = cap - len(current_blue[0])
            seen = set(current_blue[0])
            to_add = []
            for n in nums:
                if len(to_add) >= remaining:
                    break
                if 1 <= n <= 16 and n not in seen:
                    to_add.append(n); seen.add(n)
            if not to_add:
                return
            if reading_blue[0]:
                _add_blue_nums(to_add)
                return
            # 直接存入current_blue
            if not reading_blue[0]:
                reading_blue[0] = True
            _add_blue_nums(to_add)
    
    def _try_finalize_with_blue():
        """尝试用当前蓝球与等待队列中的红球配对"""
        if pending_red_groups and len(current_blue[0]) >= 1:
            # 取出最早等待的红球组
            red_group = pending_red_groups.pop(0)
            if red_group["dantuo"]:
                merged = list(set(red_group["dan"] + red_group["tuo"]))
                merged.sort()
                if len(merged) >= 2:
                    all_groups.append({
                        "red": merged,
                        "blue": sorted(current_blue[0]),
                        "times": red_group["times"]
                    })
            else:
                if len(red_group["red"]) >= 2:
                    all_groups.append({
                        "red": sorted(red_group["red"]),
                        "blue": sorted(current_blue[0]),
                        "times": red_group["times"]
                    })
            
            # 如果还有等待的红球组，保留蓝球数据用于下一组
            # 如果没有等待的红球组，才清空蓝球数据
            if pending_red_groups:
                # 还有红球等待配对，保留蓝球数据并清空reading状态
                # 这样后续的蓝球数字可以继续添加
                # 但需要重新开启读取模式
                pass  # 保留current_blue，继续累加
            else:
                current_blue[0] = []
                reading_blue[0] = False
    
    def _finalize():
        """完成当前分组（用于非队列模式或最后一组）"""
        # 先尝试用队列模式处理所有能配对的组
        while pending_red_groups and len(current_blue[0]) >= 1:
            _try_finalize_with_blue()
        
        # 如果还有未配对的蓝球但没有更多红球等待，
        # 将蓝球与最后一个current_red（如果有）配对
        if pending_red_groups:
            # 队列中还有未配对的红球，丢弃（没有蓝球）
            # 但保留current_blue用于后续可能的红球
            pass
        
        # 处理最后一组（如果有完整红蓝球）
        if in_dantuo_mode[0]:
            merged = list(set(current_red_dan[0] + current_red_tuo[0]))
            merged.sort()
            if len(merged) >= 2 and len(current_blue[0]) >= 1:
                all_groups.append({
                    "red": merged,
                    "blue": sorted(current_blue[0]),
                    "times": current_times[0]
                })
        else:
            if len(current_red[0]) >= 2 and len(current_blue[0]) >= 1:
                all_groups.append({
                    "red": sorted(current_red[0]),
                    "blue": sorted(current_blue[0]),
                    "times": current_times[0]
                })
        
        # 处理队列中剩余未配对的红球（没有蓝球，无法组成有效组）
        pending_red_groups.clear()
        
        # 重置状态
        current_red[0] = []
        current_blue[0] = []
        current_times[0] = 1
        current_red_dan[0] = []
        current_red_tuo[0] = []
        in_dantuo_mode[0] = False
        passed_tuo[0] = False
        reading_red[0] = False
        reading_blue[0] = False
    
    def _start_new_red(nums, is_dan):
        """开始新的红球读取"""
        if is_dan:
            # 开始胆拖模式 - 红胆
            in_dantuo_mode[0] = True
            passed_tuo[0] = False
            current_red_dan[0] = list(nums)
            reading_red[0] = True
            reading_blue[0] = False
        elif in_dantuo_mode[0]:
            # 胆拖模式下遇到新的红关键字（通常是"红拖"）
            passed_tuo[0] = True
            # 将数字加入拖码（如果同行有数字）
            for n in nums:
                if n not in current_red_tuo[0]:
                    current_red_tuo[0].append(n)
            reading_red[0] = True
            reading_blue[0] = False
        else:
            # 普通模式
            current_red[0] = list(nums)
            reading_red[0] = True
            reading_blue[0] = False
    
    def _add_red_nums(nums):
        """添加红球号码到当前组"""
        if in_dantuo_mode[0]:
            if passed_tuo[0]:
                # 已过"红拖"标记，加入拖码
                for n in nums:
                    if n not in current_red_tuo[0]:
                        current_red_tuo[0].append(n)
            else:
                # 还在"红胆"阶段，加入胆码
                for n in nums:
                    if n not in current_red_dan[0]:
                        current_red_dan[0].append(n)
        else:
            for n in nums:
                if n not in current_red[0]:
                    current_red[0].append(n)
    
    def _start_new_blue(nums):
        """开始新的蓝球读取"""
        # 保留已读取的红球数据
        current_blue[0] = list(nums)
        reading_blue[0] = True
        reading_red[0] = False
    
    def _add_blue_nums(nums):
        """添加蓝球号码到当前组"""
        for n in nums:
            if n not in current_blue[0]:
                current_blue[0].append(n)
    
    # ====== 辅助：将一行按红球/蓝球关键字拆分成多个逻辑行 ======
    def _split_line_by_keywords(stripped_line):
        """当一行内多次出现红球/蓝球关键字时，拆分成多个逻辑行
        
        例如："红球: 01 红球: 02 03 蓝球: 04"
        拆成：["红球: 01", "红球: 02 03", "蓝球: 04"]
        """
        # 匹配红球/红胆/红拖/蓝球等关键字的起始位置
        split_markers = []
        # 从最具体到最不具体，避免"红"先匹配
        red_kw_list = ["红球", "红胆", "红拖", "红爬", "红皿", "红昭", "红粑",
                       "纤球", "红旦", "红扡", "红施", "红肛", "红色"]
        blue_kw_list = ["蓝球", "蓝胆", "蓝琛", "篮球", "兰球", "蓝求", "篮胆", "兰胆"]
        
        # 找出所有关键字的位置
        for kw in red_kw_list + blue_kw_list:
            start = 0
            while True:
                idx = stripped_line.find(kw, start)
                if idx < 0:
                    break
                # 避免重复匹配：同一位置已记录则跳过
                if not any(m[0] == idx for m in split_markers):
                    is_red = kw in red_kw_list
                    split_markers.append((idx, len(kw), is_red))
                start = idx + 1
        # 额外匹配：单字"红:" / "蓝:"
        for pat_kw, is_red in [("红:", True), ("红：", True), ("蓝:", False), ("蓝：", False)]:
            start = 0
            while True:
                idx = stripped_line.find(pat_kw, start)
                if idx < 0:
                    break
                if not any(m[0] == idx for m in split_markers):
                    split_markers.append((idx, len(pat_kw) - 1, is_red))
                start = idx + 1
        
        if len(split_markers) <= 1:
            # 没有或只有一个关键字，不拆分
            return [stripped_line]
        
        # 按位置排序
        split_markers.sort(key=lambda x: x[0])
        
        # ===== 关键修复：合并同类型连续关键字（叠印）=====
        # 例如 "蓝球:11 蓝球:11 蓝球:11 12 13" 中3个蓝球是叠印，不应该拆成3行
        # 只保留同类型连续关键字的第一个，跳过距离<15字符的同类型关键字
        merged_markers = []
        for m in split_markers:
            if merged_markers:
                last_m = merged_markers[-1]
                # 同类型且距离很近（<15字符）→ 叠印，跳过
                if last_m[2] == m[2] and m[0] - last_m[0] < 15:
                    continue
            merged_markers.append(m)
        split_markers = merged_markers
        
        if len(split_markers) <= 1:
            # 合并后没有或只有一个关键字，不拆分
            return [stripped_line]
        
        # 拆分
        logical_lines = []
        for i, (pos, klen, _) in enumerate(split_markers):
            if i < len(split_markers) - 1:
                end = split_markers[i+1][0]
            else:
                end = len(stripped_line)
            seg = stripped_line[pos:end].strip()
            if seg:
                logical_lines.append(seg)
        return logical_lines
    
    # ====== 辅助：从整行（不管含多少中文）提取所有有效数字 ======
    def _extract_all_nums_from_line(text_line):
        """从整行文本中暴力提取所有1-33范围的数字（含OCR颠倒修正）"""
        result = []
        for nm in re.finditer(r"(\d{1,2})", text_line):
            try:
                n = int(nm.group(1))
                n = _fix_num_ocr(n)
                if 1 <= n <= 33:
                    result.append(n)
            except:
                pass
        return result
    
    # ====== 主解析循环 ======
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        
        # ===== 关键修复1：按红球/蓝球关键字拆分成多个逻辑行（解决1896d7b3合并问题） =====
        logical_lines = _split_line_by_keywords(stripped)
        
        for logic_line in logical_lines:
            logic_line = logic_line.strip()
            if not logic_line:
                continue
            
            # ===== 广告区保护：倍数行之后（in_ad_zone=True），跳过所有非关键字行 =====
            # 但"双色球"关键字可以退出广告区（标记新的一组彩票开始）
            if in_ad_zone[0]:
                # 检查是否是"双色球"开头（新组标记）
                if "双色球" in logic_line or "双色" in logic_line:
                    in_ad_zone[0] = False
                    # 不continue，继续往下检查是否同时有红球/蓝球关键字
                else:
                    # 检查是否是显式红球/蓝球关键字（退出广告区）
                    is_dan_chk, _ = _is_red_keyword(logic_line)
                    blue_chk = _is_blue_keyword(logic_line)
                    if is_dan_chk is None and not blue_chk:
                        continue  # 广告区中非关键字行，跳过
            
            # 1. 检查是否是红球/红胆/红拖等开头
            is_dan, red_nums = _is_red_keyword(logic_line)
            if is_dan is not None:
                # 遇到显式红球关键字 → 退出广告区，新的号码组开始
                in_ad_zone[0] = False
                # ===== 关键修复：组间隔离 =====
                # 新红球关键字出现时，如果蓝球已经收集了号码，说明这是【下一组彩票】的红球开始了，
                # 必须立即把之前的所有组都finalize入all_groups，清空状态，避免跨组污染
                if len(current_blue[0]) >= 1 and (pending_red_groups or len(current_red[0]) >= 1):
                    # 先把current_red保存（如果有）
                    if len(current_red[0]) >= 1 and not in_dantuo_mode[0]:
                        _save_current_red()
                    elif in_dantuo_mode[0] and (len(current_red_dan[0]) + len(current_red_tuo[0])) >= 1:
                        _save_current_red()
                    # 一次性配对所有pending+current_blue能配的
                    while pending_red_groups and len(current_blue[0]) >= 1:
                        _try_finalize_with_blue()
                    # 剩余状态强制入all_groups（包括未配对的），然后清空
                    # 注意：_finalize会清空所有状态
                    if pending_red_groups or len(current_blue[0]) >= 1 or len(current_red[0]) >= 1:
                        # 如果有剩余current_red未保存的先保存再finalize
                        # _finalize会处理剩余的
                        _finalize()
                
                # 如果已有红球队列或当前红球，先保存
                # 但在胆拖模式下遇到"红拖"时，不保存（只是胆拖内的状态转换）
                is_red_tuo_transition = (in_dantuo_mode[0] and not is_dan)
                if not is_red_tuo_transition:
                    # 如果current_red有号码，先保存（即使只有1个）
                    if in_dantuo_mode[0] or len(current_red[0]) >= 1:
                        _save_current_red()
                    # 【延迟配对原则】：不在红球关键字处立即配对，避免蓝球延续行追加时pending已空。
                    # 配对只发生在：倍数行、组间隔离（下一组红球且蓝球非空）、finalize阶段
                
                _start_new_red(red_nums, is_dan)
                # ===== 红球关键字行：限制式extra追加 =====
                # 普通模式：current_red + extra ≤ 10（避免叠印污染）
                # 胆拖模式：不限制（胆2拖N的N可以很大）
                red_extra_all = extract_red_nums(logic_line)
                seen_r2 = set(red_nums)
                extra_r2 = [n for n in red_extra_all if n not in seen_r2]
                if extra_r2:
                    total_after = len(current_red_dan[0]) + len(current_red_tuo[0]) + len(extra_r2) if in_dantuo_mode[0] else len(current_red[0]) + len(extra_r2)
                    if in_dantuo_mode[0] or total_after <= 10:
                        _append_to_last_pending_or_current(extra_r2, True)
                # ===== 红球关键字行内的倍数提取 =====
                t_in_red = extract_times(logic_line)
                if t_in_red is not None:
                    current_times[0] = t_in_red
                    if pending_red_groups:
                        pending_red_groups[-1]["times"] = current_times[0]
                continue
            
            # 2. 检查是否是蓝球开头
            blue_nums = _is_blue_keyword(logic_line)
            if blue_nums:
                # 遇到显式蓝球关键字 → 退出广告区（当前彩票的蓝球组可能刚好在倍数行之后出现，理论上不应该，但容错）
                in_ad_zone[0] = False
                # 先把已有的红球保存到等待队列
                if in_dantuo_mode[0] or len(current_red[0]) >= 1:
                    _save_current_red()
                
                _start_new_blue(blue_nums)
                # 【延迟配对原则】：不在蓝球关键字处立即配对，避免蓝球延续行（9-16）追加时pending已空
                # 配对只发生在：倍数行、组间隔离（下一组红球且蓝球非空）、finalize阶段
                
                # ===== 关键字行后半部分的蓝球号码（写在同一行的延续）=====
                # 例如 "蓝球:01 02 03 04 05 心佑 06 07 08" → 06,07,08也要提取追加
                blue_extra = extract_blue_nums(logic_line)
                seen_b = set(blue_nums)
                extra_b = [n for n in blue_extra if n not in seen_b]
                if extra_b:
                    _append_to_last_pending_or_current(extra_b, False)
                # ===== 关键字行内的倍数提取 =====
                # 例如 "蓝球:11 12 13 [1倍]" → 倍数1意味着蓝球已收齐
                # 关闭reading_blue + 配对 + 进入广告区（后续行不再追加）
                t_in_blue = extract_times(logic_line)
                if t_in_blue is not None:
                    current_times[0] = t_in_blue
                    if pending_red_groups:
                        pending_red_groups[-1]["times"] = current_times[0]
                    # 蓝球已收齐：关闭reading + 配对 + 进入广告区
                    reading_red[0] = False
                    reading_blue[0] = False
                    in_ad_zone[0] = True
                    while pending_red_groups and len(current_blue[0]) >= 1:
                        _try_finalize_with_blue()
                continue
            
            # 3. 检查是否是纯倍数行（没有红球/蓝球关键字）
            # 关键修复：倍数行检查移到关键字之后，避免"蓝球:11 12 [1倍]"被倍数行提前拦截
            t = extract_times(logic_line)
            if t is not None:
                current_times[0] = t
                if pending_red_groups:
                    pending_red_groups[-1]["times"] = current_times[0]
                # ===== 号码区结束标记 =====
                reading_red[0] = False
                reading_blue[0] = False
                in_ad_zone[0] = True  # 倍数行之后进入广告区
                # 有倍数行出现，如果红球蓝球都齐全了，尝试配对（所有能配的都配）
                while pending_red_groups and len(current_blue[0]) >= 1:
                    _try_finalize_with_blue()
                continue
            
            # ===== 暴力提取整行所有数字 + 追加到pending队列或current =====
            # 4. 从包含数字的行提取号码 - 支持追加到pending最后一组
            if in_ad_zone[0]:
                continue
            has_digit = re.search(r"\d", logic_line)
            if has_digit:
                # 提取红球范围的号码（1-33）
                red_extracted = extract_red_nums(logic_line)
                # 提取蓝球范围的号码（1-16）
                blue_extracted = extract_blue_nums(logic_line)
                
                # 根据上下文决定是红球延续还是蓝球延续
                if reading_blue[0] or (len(current_blue[0]) > 0 and not reading_red[0] and len(blue_extracted) >= len(red_extracted)):
                    # 在蓝球读取模式，或：已有current_blue且没有在reading_red且蓝球范围数字更多
                    if blue_extracted:
                        _append_to_last_pending_or_current(blue_extracted, False)
                        # 如果还有红号范围数字且pending队列未满（可能是行内夹杂红球但主要是蓝球延续），跳过
                        continue
                if reading_red[0] or pending_red_groups or len(current_red[0]) > 0:
                    if red_extracted:
                        _append_to_last_pending_or_current(red_extracted, True)
                        continue
                # 兜底：不在任何模式但有数字 - 如果蓝球更多且蓝球范围合法（可能是关键字漏掉的蓝球行）
                if len(blue_extracted) >= 3 and len(current_blue[0]) == 0 and not reading_blue[0]:
                    _append_to_last_pending_or_current(blue_extracted, False)
                    continue
                # 兜底：有大量红号范围数字但没识别到红球关键字
                if len(red_extracted) >= 5 and not pending_red_groups and len(current_red[0]) == 0:
                    # 当作隐式红球组开启（交给后续蓝球配对）
                    reading_red[0] = True
                    _add_red_nums(red_extracted)
                    continue
                # 最后：还没匹配到情况，至少把数字追加到合理位置
                if red_extracted and (pending_red_groups or reading_red[0] or len(current_red[0]) > 0):
                    _append_to_last_pending_or_current(red_extracted, True)
                    continue
                if blue_extracted and len(current_blue[0]) > 0:
                    _append_to_last_pending_or_current(blue_extracted, False)
                    continue
    
    # 处理最后一组
    _finalize()
    
    if debug:
        print(f"  解析到 {len(all_groups)} 组复式")
        for g in all_groups:
            print(f"    红球: {g['red']}, 蓝球: {g['blue']}, 倍数: {g['times']}")
    
    return all_groups

# ===================== 4. 批量遍历文件夹所有图片，汇总全部号码 =====================
def batch_parse_images(folder_path, debug=False):
    total_groups = []
    # 遍历jpg/png图片
    for file_name in os.listdir(folder_path):
        if file_name.lower().endswith((".jpg", ".png", ".jpeg")):
            img_full_path = os.path.join(folder_path, file_name)
            if debug:
                print(f"\n{'='*60}")
                print(f"处理图片: {file_name}")
            groups = parse_lottery_image(img_full_path, debug=debug)
            for g in groups:
                g["img_name"] = file_name
                total_groups.append(g)
            print(f"【{file_name}】识别到 {len(groups)} 组复式")
            if debug and groups:
                for g in groups:
                    print(f"  红球: {g['red']}, 蓝球: {g['blue']}, 倍数: {g['times']}")
    print(f"\n全部图片解析完成，总计 {len(total_groups)} 组彩票复式")
    return total_groups

# ===================== 5. 统计红球、蓝球出现频次（加权倍数） =====================
def calc_hot_cold(all_groups):
    red_count = defaultdict(int)
    blue_count = defaultdict(int)
    for item in all_groups:
        mul = item["times"]
        # 统计红球
        for r in item["red"]:
            red_count[r] += mul
        # 统计蓝球
        for b in item["blue"]:
            blue_count[b] += mul
    # 按出现次数降序排序（热号在前）
    sorted_red = sorted(red_count.items(), key=lambda x: x[1], reverse=True)
    sorted_blue = sorted(blue_count.items(), key=lambda x: x[1], reverse=True)
    return sorted_red, sorted_blue

# ===================== 6. 根据热度推演娱乐推荐号码 =====================
def get_recommend(sorted_red, sorted_blue):
    # 取热度前12个红球作为备选池
    hot_red_pool = [num for num, cnt in sorted_red[:12]]
    # 取热度前5个蓝球作为备选池
    hot_blue_pool = [num for num, cnt in sorted_blue[:5]]

    print("\n========== 热度统计结果 ==========")
    print("红球热度排序（数字:出现次数）：")
    for num, cnt in sorted_red:
        print(f"{num:02d} : {cnt}次")

    print("\n蓝球热度排序（数字:出现次数）：")
    for num, cnt in sorted_blue:
        print(f"{num:02d} : {cnt}次")

    print("\n========== 娱乐推荐号码池 ==========")
    print(f"高频红球池(任选6个组合)：{hot_red_pool}")
    print(f"高频蓝球池(任选1个搭配)：{hot_blue_pool}")
    return hot_red_pool, hot_blue_pool

# ===================== 7. 导出所有彩票分组到Excel备查 =====================
def export_to_excel(all_groups, save_name=None):
    """导出双色球号码汇总到Excel

    输出格式：
      - 三列：序号、红球、蓝球
      - 号码以空格分隔，两位数字带前导零（如 04 11 12 13 15 17 28）
      - 表头加粗居中，数据居中对齐，带边框，列宽自适应
    """
    if save_name is None:
        save_name = LOTTERY_SUMMARY_PATH
    save_name = os.path.abspath(save_name)

    # 构建数据：序号、红球、蓝球三列
    data_list = []
    for idx, g in enumerate(all_groups, start=1):
        data_list.append({
            "序号": idx,
            "红球": " ".join([f"{x:02d}" for x in g["red"]]),
            "蓝球": " ".join([f"{x:02d}" for x in g["blue"]]),
        })
    df = pd.DataFrame(data_list)

    # 写入Excel
    df.to_excel(save_name, index=False, engine="openpyxl")

    # ===== 样式优化 =====
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, Border, Side

        wb = load_workbook(save_name)
        ws = wb.active

        # 定义样式
        header_font = Font(name="微软雅黑", size=12, bold=True)
        header_align = Alignment(horizontal="center", vertical="center")
        data_font = Font(name="微软雅黑", size=11)
        data_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        # 表头样式（第1行）
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border

        # 数据行样式
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border

        # 列宽设置
        ws.column_dimensions["A"].width = 8    # 序号列
        ws.column_dimensions["B"].width = 30   # 红球列
        ws.column_dimensions["C"].width = 16   # 蓝球列

        # 行高
        ws.row_dimensions[1].height = 24
        for r in range(2, ws.max_row + 1):
            ws.row_dimensions[r].height = 22

        wb.save(save_name)
    except ImportError:
        pass
    except Exception as e:
        print(f"  [警告] Excel样式优化失败（不影响数据）: {e}")

    print(f"\n所有分组已导出至：{save_name}")
    print(f"  共 {len(data_list)} 组，格式：序号 | 红球 | 蓝球（空格分隔，前导零）")

# ===================== 主程序入口 =====================
if __name__ == "__main__":
    # # 普通运行
    # python3 scripts/lottery.py
    # # 调试模式（查看详细 OCR 识别结果）
    # python3 scripts/lottery.py --debug
    # # 指定图片目录
    # python3 scripts/lottery.py --folder ./your_images
    # 检查是否启用调试模式
    import argparse
    parser = argparse.ArgumentParser(description='双色球图片OCR识别')
    parser.add_argument('--debug', '-d', action='store_true', help='启用调试模式，显示详细OCR输出')
    parser.add_argument('--folder', '-f', default='../lottery_img', help='图片文件夹路径')
    args = parser.parse_args()
    
    # 1. 设置图片文件夹路径
    IMG_FOLDER = args.folder

    # 2. 批量解析所有图片
    all_lottery_groups = batch_parse_images(IMG_FOLDER, debug=args.debug)

    # 3. 导出全部号码到Excel
    export_to_excel(all_lottery_groups)
    print("---------------end")
    # 4. 计算冷热号频次
    # red_hot_sort, blue_hot_sort = calc_hot_cold(all_lottery_groups)

    # 5. 输出热度并生成推荐号码池
    # red_pool, blue_pool = get_recommend(red_hot_sort, blue_hot_sort)